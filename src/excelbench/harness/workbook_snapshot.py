"""Semantic workbook snapshot extraction for workbook/package diffs."""

from __future__ import annotations

import hashlib
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from zipfile import ZipFile

import openpyxl

JSONDict = dict[str, Any]

_VOLATILE_PARTS = {
    "docProps/core.xml",
}


@dataclass(frozen=True)
class WorkbookSnapshot:
    """Normalized workbook semantics plus selected OOXML package metadata."""

    workbook: str
    categories: dict[str, Any]

    def to_json_dict(self) -> JSONDict:
        return {"workbook": self.workbook, "categories": self.categories}


def snapshot_workbook(path: Path) -> WorkbookSnapshot:
    """Build a deterministic semantic snapshot for an OOXML workbook."""
    path = Path(path)
    workbook = openpyxl.load_workbook(path, data_only=False)
    try:
        categories: dict[str, Any] = {
            "sheets": _snapshot_sheets(workbook),
            "cells": _snapshot_cells(workbook),
            "formats": _snapshot_formats(workbook),
            "merges": _snapshot_merges(workbook),
            "validations": _snapshot_validations(workbook),
            "hyperlinks": _snapshot_hyperlinks(workbook),
            "comments": _snapshot_comments(workbook),
            "freeze_panes": _snapshot_freeze_panes(workbook),
            "named_ranges": _snapshot_named_ranges(workbook),
            "tables": _snapshot_tables(workbook),
            "images": _snapshot_images(workbook),
            "pivot_metadata": _snapshot_pivot_metadata(path),
            "package_parts": _snapshot_package_parts(path),
        }
    finally:
        workbook.close()
    return WorkbookSnapshot(workbook=str(path), categories=_json_stable(categories))


def write_snapshot(snapshot: WorkbookSnapshot, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(snapshot.to_json_dict(), indent=2, sort_keys=True) + "\n")


def _snapshot_sheets(workbook: Any) -> JSONDict:
    return {
        "names": list(workbook.sheetnames),
        "states": {name: workbook[name].sheet_state for name in workbook.sheetnames},
    }


def _snapshot_cells(workbook: Any) -> JSONDict:
    cells: JSONDict = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheet_cells: JSONDict = {}
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if (
                    value is None
                    and not cell.has_style
                    and cell.comment is None
                    and cell.hyperlink is None
                ):
                    continue
                sheet_cells[cell.coordinate] = {
                    "value": _normalize_value(value),
                    "data_type": cell.data_type,
                }
        cells[sheet_name] = sheet_cells
    return cells


def _snapshot_formats(workbook: Any) -> JSONDict:
    formats: JSONDict = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheet_formats: JSONDict = {}
        for row in sheet.iter_rows():
            for cell in row:
                if not cell.has_style:
                    continue
                fmt = _cell_format(cell)
                if fmt:
                    sheet_formats[cell.coordinate] = fmt
        formats[sheet_name] = sheet_formats
    return formats


def _snapshot_merges(workbook: Any) -> JSONDict:
    return {
        name: sorted(str(rng) for rng in workbook[name].merged_cells.ranges)
        for name in workbook.sheetnames
    }


def _snapshot_validations(workbook: Any) -> JSONDict:
    out: JSONDict = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        validations = []
        for dv in getattr(sheet.data_validations, "dataValidation", []):
            validations.append(
                {
                    "sqref": str(dv.sqref),
                    "type": dv.type,
                    "operator": dv.operator,
                    "formula1": dv.formula1,
                    "formula2": dv.formula2,
                    "allow_blank": bool(dv.allow_blank),
                }
            )
        out[sheet_name] = sorted(validations, key=lambda item: (item["sqref"], item["type"] or ""))
    return out


def _snapshot_hyperlinks(workbook: Any) -> JSONDict:
    out: JSONDict = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        links: JSONDict = {}
        for row in sheet.iter_rows():
            for cell in row:
                link = cell.hyperlink
                if link is None:
                    continue
                links[cell.coordinate] = {
                    "target": link.target,
                    "location": link.location,
                    "display": link.display,
                    "tooltip": link.tooltip,
                }
        out[sheet_name] = links
    return out


def _snapshot_comments(workbook: Any) -> JSONDict:
    out: JSONDict = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        comments: JSONDict = {}
        for row in sheet.iter_rows():
            for cell in row:
                if cell.comment is not None:
                    comments[cell.coordinate] = {
                        "text": cell.comment.text,
                        "author": cell.comment.author,
                    }
        out[sheet_name] = comments
    return out


def _snapshot_freeze_panes(workbook: Any) -> JSONDict:
    return {name: str(workbook[name].freeze_panes or "") for name in workbook.sheetnames}


def _snapshot_named_ranges(workbook: Any) -> JSONDict:
    ranges: JSONDict = {}
    for name, defined_name in workbook.defined_names.items():
        ranges[name] = {
            "attr_text": defined_name.attr_text,
            "local_sheet_id": defined_name.localSheetId,
            "hidden": bool(defined_name.hidden),
        }
    return ranges


def _snapshot_tables(workbook: Any) -> JSONDict:
    out: JSONDict = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        tables: JSONDict = {}
        for name, table in sheet.tables.items():
            table_obj = sheet.tables[name]
            tables[name] = {
                "ref": table_obj.ref,
                "display_name": table_obj.displayName,
                "totals_row_shown": bool(table_obj.totalsRowShown),
                "table_style": getattr(table_obj.tableStyleInfo, "name", None),
            }
        out[sheet_name] = tables
    return out


def _snapshot_images(workbook: Any) -> JSONDict:
    out: JSONDict = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        images = []
        for image in getattr(sheet, "_images", []):
            anchor = getattr(image, "anchor", None)
            images.append(
                {
                    "format": getattr(image, "format", None),
                    "width": getattr(image, "width", None),
                    "height": getattr(image, "height", None),
                    "anchor": _anchor_to_text(anchor),
                }
            )
        out[sheet_name] = images
    return out


def _snapshot_pivot_metadata(path: Path) -> JSONDict:
    with ZipFile(path) as workbook_zip:
        parts = sorted(
            name
            for name in workbook_zip.namelist()
            if "pivot" in name.lower() or "pivotcache" in name.lower()
        )
    return {"parts": parts, "count": len(parts)}


def _snapshot_package_parts(path: Path) -> JSONDict:
    parts: JSONDict = {}
    with ZipFile(path) as workbook_zip:
        for info in workbook_zip.infolist():
            if info.filename in _VOLATILE_PARTS or info.is_dir():
                continue
            payload = workbook_zip.read(info.filename)
            parts[info.filename] = {
                "size": len(payload),
                "sha256": hashlib.sha256(payload).hexdigest(),
            }
    return parts


def _cell_format(cell: Any) -> JSONDict:
    font = cell.font
    fill = cell.fill
    alignment = cell.alignment
    return _drop_empty(
        {
            "number_format": cell.number_format,
            "font": _drop_empty(
                {
                    "name": font.name,
                    "size": font.sz,
                    "bold": bool(font.b),
                    "italic": bool(font.i),
                    "underline": font.u,
                    "strike": bool(font.strike),
                    "color": _color_to_text(font.color),
                }
            ),
            "fill": _drop_empty(
                {
                    "type": fill.fill_type,
                    "fg_color": _color_to_text(fill.fgColor),
                    "bg_color": _color_to_text(fill.bgColor),
                }
            ),
            "alignment": _drop_empty(
                {
                    "horizontal": alignment.horizontal,
                    "vertical": alignment.vertical,
                    "wrap_text": alignment.wrap_text,
                    "text_rotation": alignment.textRotation,
                    "indent": alignment.indent,
                }
            ),
        }
    )


def _normalize_value(value: Any) -> Any:
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def _color_to_text(color: Any) -> str | None:
    if color is None:
        return None
    color_type = getattr(color, "type", None)
    if color_type == "rgb":
        return str(getattr(color, "rgb", "") or "").upper() or None
    if color_type == "indexed":
        return f"indexed:{getattr(color, 'indexed', '')}"
    if color_type == "theme":
        return f"theme:{getattr(color, 'theme', '')}:{getattr(color, 'tint', 0)}"
    return str(color)


def _anchor_to_text(anchor: Any) -> str:
    if isinstance(anchor, str):
        return anchor
    marker = getattr(anchor, "_from", None)
    if marker is None:
        return str(anchor)
    row = int(getattr(marker, "row", 0)) + 1
    col = int(getattr(marker, "col", 0)) + 1
    return f"R{row}C{col}"


def _drop_empty(payload: JSONDict) -> JSONDict:
    return {key: value for key, value in payload.items() if value not in (None, "", {}, [])}


def _json_stable(value: Any) -> Any:
    return json.loads(json.dumps(value, sort_keys=True, default=str))
