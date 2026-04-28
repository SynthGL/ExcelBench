"""Validate external-oracle fixture preservation through WolfXL."""

from __future__ import annotations

import json
import shutil
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from pathlib import Path
from typing import Any, cast
from zipfile import ZipFile

JSONDict = dict[str, Any]


@dataclass(frozen=True)
class WolfXLFixtureValidation:
    """WolfXL read/modify-save preservation result for one fixture."""

    fixture_id: str
    source_workbook: Path
    modified_workbook: Path
    expected_parts: tuple[str, ...]
    readback_probes: tuple[JSONDict, ...]
    missing_parts_after_save: tuple[str, ...]
    readback_failures: tuple[str, ...]
    read_passed: bool
    modify_save_passed: bool
    marker_passed: bool
    error: str | None = None

    @property
    def passed(self) -> bool:
        """Return whether read, modify-save, marker, and part checks passed."""
        return (
            self.read_passed
            and self.modify_save_passed
            and self.marker_passed
            and not self.missing_parts_after_save
            and not self.readback_failures
            and self.error is None
        )

    def to_json_dict(self, fixture_root: Path) -> JSONDict:
        """Convert the validation result to a JSON manifest row."""
        return {
            "fixture_id": self.fixture_id,
            "source_workbook": _display_path(self.source_workbook, fixture_root),
            "modified_workbook": _display_path(self.modified_workbook, fixture_root),
            "expected_parts": list(self.expected_parts),
            "readback_probes": list(self.readback_probes),
            "missing_parts_after_save": list(self.missing_parts_after_save),
            "readback_failures": list(self.readback_failures),
            "read_passed": self.read_passed,
            "modify_save_passed": self.modify_save_passed,
            "marker_passed": self.marker_passed,
            "passed": self.passed,
            "error": self.error,
        }


def validate_wolfxl_external_fixture_pack(
    fixture_root: Path,
    *,
    output_dir: Path | None = None,
    marker_cell: str = "J1",
    marker_value: str = "wolfxl_modify_smoke",
) -> list[WolfXLFixtureValidation]:
    """Run WolfXL read and in-place modify-save checks over a fixture pack.

    Args:
        fixture_root: Directory containing `manifest.json` and generated
            workbook files.
        output_dir: Directory for WolfXL-modified copies. Defaults to
            `fixture_root / "wolfxl-modified"`.
        marker_cell: Cell written on the first sheet during modify-save.
        marker_value: Value expected in `marker_cell` after save.

    Returns:
        One validation result per manifest fixture entry. The function also
        writes `wolfxl-validation.json` under `fixture_root`.
    """
    fixture_root = fixture_root.resolve()
    output_dir = (output_dir or fixture_root / "wolfxl-modified").resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    manifest = _load_manifest(fixture_root)
    results: list[WolfXLFixtureValidation] = []

    for entry in manifest.get("fixtures", []):
        fixture_id = str(entry.get("fixture_id", "unknown"))
        source_workbook = fixture_root / str(entry["workbook"])
        expected_parts = tuple(str(part) for part in entry.get("expected_parts", []))
        readback_probes = tuple(
            cast(JSONDict, probe) for probe in entry.get("readback_probes", [])
        )
        modified_workbook = output_dir / source_workbook.name
        results.append(
            _validate_one_fixture(
                fixture_id=fixture_id,
                source_workbook=source_workbook,
                modified_workbook=modified_workbook,
                expected_parts=expected_parts,
                readback_probes=readback_probes,
                marker_cell=marker_cell,
                marker_value=marker_value,
            )
        )

    payload = {
        "fixture_root": str(fixture_root),
        "results": [result.to_json_dict(fixture_root) for result in results],
        "passed": all(result.passed for result in results),
    }
    (fixture_root / "wolfxl-validation.json").write_text(
        json.dumps(payload, indent=2, sort_keys=True) + "\n"
    )
    return results


def _validate_one_fixture(
    *,
    fixture_id: str,
    source_workbook: Path,
    modified_workbook: Path,
    expected_parts: tuple[str, ...],
    readback_probes: tuple[JSONDict, ...],
    marker_cell: str,
    marker_value: str,
) -> WolfXLFixtureValidation:
    try:
        import openpyxl
        import wolfxl

        read_passed = False
        marker_passed = False
        shutil.copy2(source_workbook, modified_workbook)
        before_parts = _zip_part_names(source_workbook)
        tracked_parts = tuple(part for part in expected_parts if part in before_parts)

        workbook = wolfxl.load_workbook(modified_workbook, modify=True)
        sheet_name = workbook.sheetnames[0]
        workbook[sheet_name][marker_cell] = marker_value
        read_passed = workbook[sheet_name]["A1"].value is not None
        workbook.save(modified_workbook)
        workbook.close()

        after_parts = _zip_part_names(modified_workbook)
        missing_parts = tuple(part for part in tracked_parts if part not in after_parts)
        roundtrip = openpyxl.load_workbook(modified_workbook)
        try:
            marker_passed = roundtrip[sheet_name][marker_cell].value == marker_value
            readback_failures = _run_readback_probes(
                roundtrip,
                modified_workbook,
                readback_probes,
            )
        finally:
            roundtrip.close()
        return WolfXLFixtureValidation(
            fixture_id=fixture_id,
            source_workbook=source_workbook,
            modified_workbook=modified_workbook,
            expected_parts=expected_parts,
            readback_probes=readback_probes,
            missing_parts_after_save=missing_parts,
            readback_failures=readback_failures,
            read_passed=read_passed,
            modify_save_passed=True,
            marker_passed=marker_passed,
        )
    except Exception as exc:
        return WolfXLFixtureValidation(
            fixture_id=fixture_id,
            source_workbook=source_workbook,
            modified_workbook=modified_workbook,
            expected_parts=expected_parts,
            readback_probes=readback_probes,
            missing_parts_after_save=(),
            readback_failures=(),
            read_passed=False,
            modify_save_passed=False,
            marker_passed=False,
            error=f"{type(exc).__name__}: {exc}",
        )


def _load_manifest(fixture_root: Path) -> JSONDict:
    manifest_path = fixture_root / "manifest.json"
    if not manifest_path.exists():
        raise FileNotFoundError(f"External fixture manifest not found: {manifest_path}")
    return cast(JSONDict, json.loads(manifest_path.read_text()))


def _zip_part_names(path: Path) -> set[str]:
    with ZipFile(path) as workbook_zip:
        return set(workbook_zip.namelist())


def _run_readback_probes(
    workbook: Any,
    workbook_path: Path,
    probes: tuple[JSONDict, ...],
) -> tuple[str, ...]:
    failures: list[str] = []
    for index, probe in enumerate(probes, start=1):
        kind = str(probe.get("kind", ""))
        label = str(probe.get("label") or f"probe {index}")
        try:
            if kind == "cell_value":
                _check_cell_value(workbook, probe)
            elif kind == "cell_formula":
                _check_cell_formula(workbook, probe)
            elif kind == "cell_style":
                _check_cell_style(workbook, probe)
            elif kind == "conditional_formatting":
                _check_conditional_formatting(workbook, probe)
            elif kind == "comment_text":
                _check_comment_text(workbook, probe)
            elif kind == "hyperlink_target":
                _check_hyperlink_target(workbook, probe)
            elif kind == "data_validation":
                _check_data_validation(workbook, probe)
            elif kind == "merged_range":
                _check_merged_range(workbook, probe)
            elif kind == "table_metadata":
                _check_table_metadata(workbook, probe)
            elif kind == "relationship_target":
                _check_relationship_target(workbook_path, probe)
            elif kind == "zip_contains":
                _check_zip_contains(workbook_path, probe)
            else:
                raise AssertionError(f"unknown probe kind {kind!r}")
        except Exception as exc:
            failures.append(f"{label}: {type(exc).__name__}: {exc}")
    return tuple(failures)


def _check_cell_value(workbook: Any, probe: JSONDict) -> None:
    value = workbook[str(probe["sheet"])][str(probe["cell"])].value
    expected = probe.get("expected")
    if value != expected:
        raise AssertionError(f"expected {expected!r}, got {value!r}")


def _check_cell_formula(workbook: Any, probe: JSONDict) -> None:
    value = workbook[str(probe["sheet"])][str(probe["cell"])].value
    expected = str(probe["expected"])
    if value != expected:
        raise AssertionError(f"expected formula {expected!r}, got {value!r}")


def _check_cell_style(workbook: Any, probe: JSONDict) -> None:
    cell = workbook[str(probe["sheet"])][str(probe["cell"])]
    expected = cast(JSONDict, probe.get("expected", {}))
    checks = {
        "number_format": cell.number_format,
        "font_bold": cell.font.bold,
        "font_italic": cell.font.italic,
    }
    for key, expected_value in expected.items():
        actual = checks.get(str(key))
        if actual != expected_value:
            raise AssertionError(f"expected {key}={expected_value!r}, got {actual!r}")


def _check_comment_text(workbook: Any, probe: JSONDict) -> None:
    comment = workbook[str(probe["sheet"])][str(probe["cell"])].comment
    expected = str(probe["contains"])
    if comment is None or expected not in comment.text:
        raise AssertionError(f"expected comment containing {expected!r}")


def _check_conditional_formatting(workbook: Any, probe: JSONDict) -> None:
    sheet = workbook[str(probe["sheet"])]
    expected_sqref = str(probe["sqref"])
    expected_type = str(probe["type"])
    expected_priority = probe.get("priority")
    expected_operator = probe.get("operator")
    expected_formula = probe.get("formula")
    for formatting in sheet.conditional_formatting:
        if str(formatting.sqref) != expected_sqref:
            continue
        for rule in formatting.rules:
            if rule.type != expected_type:
                continue
            if expected_priority is not None and rule.priority != expected_priority:
                raise AssertionError(
                    f"expected priority {expected_priority!r}, got {rule.priority!r}"
                )
            if expected_operator is not None and rule.operator != expected_operator:
                raise AssertionError(
                    f"expected operator {expected_operator!r}, got {rule.operator!r}"
                )
            if expected_formula is not None and list(rule.formula or []) != [expected_formula]:
                raise AssertionError(
                    f"expected formula {[expected_formula]!r}, got {list(rule.formula or [])!r}"
                )
            return
    raise AssertionError(f"expected {expected_type!r} conditional format on {expected_sqref!r}")


def _check_hyperlink_target(workbook: Any, probe: JSONDict) -> None:
    hyperlink = workbook[str(probe["sheet"])][str(probe["cell"])].hyperlink
    expected = str(probe["target"])
    actual = hyperlink.target if hyperlink is not None else None
    if actual != expected:
        raise AssertionError(f"expected hyperlink target {expected!r}, got {actual!r}")


def _check_data_validation(workbook: Any, probe: JSONDict) -> None:
    sheet = workbook[str(probe["sheet"])]
    expected_cell = str(probe["cell"])
    expected_type = probe.get("type")
    expected_formula1 = probe.get("formula1")
    for validation in sheet.data_validations.dataValidation:
        sqref = str(validation.sqref)
        cells = {part.strip() for part in sqref.split()}
        if expected_cell not in cells:
            continue
        if expected_type is not None and validation.type != expected_type:
            raise AssertionError(
                f"expected validation type {expected_type!r}, got {validation.type!r}"
            )
        if expected_formula1 is not None and validation.formula1 != expected_formula1:
            raise AssertionError(
                f"expected validation formula1 {expected_formula1!r}, got {validation.formula1!r}"
            )
        return
    raise AssertionError(f"expected data validation covering {expected_cell!r}")


def _check_merged_range(workbook: Any, probe: JSONDict) -> None:
    ranges = {str(cell_range) for cell_range in workbook[str(probe["sheet"])].merged_cells.ranges}
    expected = str(probe["range"])
    if expected not in ranges:
        raise AssertionError(f"expected merged range {expected!r}, got {sorted(ranges)}")


def _check_table_metadata(workbook: Any, probe: JSONDict) -> None:
    sheet = workbook[str(probe["sheet"])]
    table_name = str(probe["name"])
    if table_name not in sheet.tables:
        raise AssertionError(f"expected table {table_name!r}, got {sorted(sheet.tables)}")
    table = sheet.tables[table_name]
    expected_ref = probe.get("ref")
    if expected_ref is not None and table.ref != expected_ref:
        raise AssertionError(f"expected table ref {expected_ref!r}, got {table.ref!r}")
    expected_style = probe.get("style")
    style = getattr(table, "tableStyleInfo", None)
    actual_style = style.name if style is not None else None
    if expected_style is not None and actual_style != expected_style:
        raise AssertionError(f"expected table style {expected_style!r}, got {actual_style!r}")


def _check_relationship_target(workbook_path: Path, probe: JSONDict) -> None:
    part = str(probe["part"])
    expected_target = str(probe["target"])
    expected_type_contains = probe.get("type_contains")
    with ZipFile(workbook_path) as workbook_zip:
        root = ET.fromstring(workbook_zip.read(part))
    for relationship in root:
        target = relationship.attrib.get("Target")
        rel_type = relationship.attrib.get("Type", "")
        if target != expected_target:
            continue
        if expected_type_contains is not None and str(expected_type_contains) not in rel_type:
            raise AssertionError(
                "expected relationship type containing "
                f"{expected_type_contains!r}, got {rel_type!r}"
            )
        return
    raise AssertionError(f"expected relationship target {expected_target!r} in {part}")


def _check_zip_contains(workbook_path: Path, probe: JSONDict) -> None:
    part = str(probe["part"])
    expected = str(probe["contains"])
    with ZipFile(workbook_path) as workbook_zip:
        text = workbook_zip.read(part).decode("utf-8", errors="ignore")
    if expected not in text:
        raise AssertionError(f"expected {part} to contain {expected!r}")


def _display_path(path: Path, root: Path) -> str:
    try:
        return str(path.relative_to(root))
    except ValueError:
        return str(path)
