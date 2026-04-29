"""Openpyxl-style compatibility context cases."""

from __future__ import annotations

import json
import platform
from collections.abc import Callable
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Any

from excelbench.harness.semantic_diff import write_diff_artifacts

JSONDict = dict[str, Any]
CaseFunc = Callable[[Any, Path], Any]


@dataclass(frozen=True)
class CompatibilityCase:
    case_id: str
    api_surface: str
    description: str
    run: CaseFunc


@dataclass(frozen=True)
class CompatibilityResult:
    adapter: str
    case_id: str
    api_surface: str
    passed: bool
    skipped: bool
    return_match: bool
    snapshot_match: bool
    delta_count: int
    diff_dir: str | None = None
    error: str | None = None

    def to_json_dict(self) -> JSONDict:
        return {
            "adapter": self.adapter,
            "case_id": self.case_id,
            "api_surface": self.api_surface,
            "passed": self.passed,
            "skipped": self.skipped,
            "return_match": self.return_match,
            "snapshot_match": self.snapshot_match,
            "delta_count": self.delta_count,
            "diff_dir": self.diff_dir,
            "error": self.error,
        }


def compatibility_cases() -> list[CompatibilityCase]:
    """Return the initial broad openpyxl-style compatibility case set."""
    return [
        CompatibilityCase("create_string", "cells", "write a string", _case_create_string),
        CompatibilityCase("create_number", "cells", "write a number", _case_create_number),
        CompatibilityCase("create_formula", "formulas", "write a formula", _case_create_formula),
        CompatibilityCase("font_bold", "styles", "set bold font", _case_font_bold),
        CompatibilityCase("font_color", "styles", "set font color", _case_font_color),
        CompatibilityCase("fill_color", "styles", "set fill color", _case_fill_color),
        CompatibilityCase("number_format", "styles", "set number format", _case_number_format),
        CompatibilityCase("alignment_wrap", "styles", "set wrap alignment", _case_alignment_wrap),
        CompatibilityCase("comment", "comments", "add a comment", _case_comment),
        CompatibilityCase("hyperlink", "hyperlinks", "add a hyperlink", _case_hyperlink),
        CompatibilityCase(
            "validation",
            "data_validations",
            "add list validation",
            _case_validation,
        ),
        CompatibilityCase("merge", "merges", "merge cells", _case_merge),
        CompatibilityCase("unmerge", "merges", "merge then unmerge cells", _case_unmerge),
        CompatibilityCase("freeze_panes", "freeze_panes", "freeze panes", _case_freeze),
        CompatibilityCase("named_range", "named_ranges", "add defined name", _case_named_range),
        CompatibilityCase("table", "tables", "add table", _case_table),
        CompatibilityCase("sheet_create", "sheets", "create a second sheet", _case_sheet_create),
        CompatibilityCase("sheet_rename", "sheets", "rename a sheet", _case_sheet_rename),
        CompatibilityCase("row_height", "dimensions", "set row height", _case_row_height),
        CompatibilityCase("column_width", "dimensions", "set column width", _case_column_width),
    ]


def run_compatibility_context(
    output_dir: Path,
    *,
    adapter_names: list[str] | None = None,
) -> JSONDict:
    """Run compatibility snippets against openpyxl and selected compatible adapters."""
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    adapters = _resolve_compatible_modules(adapter_names)
    cases = compatibility_cases()
    results: list[CompatibilityResult] = []

    with TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        for case in cases:
            reference_path = tmp_path / f"{case.case_id}.openpyxl.xlsx"
            reference_return = case.run(_openpyxl_module(), reference_path)
            for adapter_name, module in adapters.items():
                if adapter_name == "openpyxl":
                    continue
                target_path = tmp_path / f"{case.case_id}.{adapter_name}.xlsx"
                try:
                    target_return = case.run(module, target_path)
                    diff_dir = output_dir / "diffs" / adapter_name / case.case_id
                    diff = write_diff_artifacts(reference_path, target_path, diff_dir)
                    return_match = _stable_json(reference_return) == _stable_json(target_return)
                    results.append(
                        CompatibilityResult(
                            adapter=adapter_name,
                            case_id=case.case_id,
                            api_surface=case.api_surface,
                            passed=return_match and diff.passed,
                            skipped=False,
                            return_match=return_match,
                            snapshot_match=diff.passed,
                            delta_count=len(diff.deltas),
                            diff_dir=str(diff_dir),
                        )
                    )
                except Exception as exc:
                    results.append(
                        CompatibilityResult(
                            adapter=adapter_name,
                            case_id=case.case_id,
                            api_surface=case.api_surface,
                            passed=False,
                            skipped=False,
                            return_match=False,
                            snapshot_match=False,
                            delta_count=0,
                            error=f"{type(exc).__name__}: {exc}",
                        )
                    )

    requested = adapter_names or sorted(adapters)
    skipped = [
        CompatibilityResult(
            adapter=name,
            case_id="all",
            api_surface="adapter",
            passed=False,
            skipped=True,
            return_match=False,
            snapshot_match=False,
            delta_count=0,
            error="Adapter is not openpyxl-compatible for snippet execution.",
        )
        for name in requested
        if name not in adapters and name != "openpyxl"
    ]
    all_results = [*results, *skipped]
    payload: JSONDict = {
        "generated_at": datetime.now(UTC).isoformat(),
        "platform": f"{platform.system()}-{platform.machine()}",
        "cases": [
            {
                "case_id": case.case_id,
                "api_surface": case.api_surface,
                "description": case.description,
            }
            for case in cases
        ],
        "results": [result.to_json_dict() for result in all_results],
        "passed": all(result.passed or result.skipped for result in all_results),
    }
    (output_dir / "compatibility.json").write_text(
        json.dumps(payload, indent=2, sort_keys=True) + "\n"
    )
    (output_dir / "COMPATIBILITY.md").write_text(_render_compatibility_markdown(payload))
    (output_dir / "CONTEXT.md").write_text(_render_compatibility_context(requested))
    return payload


def _resolve_compatible_modules(adapter_names: list[str] | None) -> dict[str, Any]:
    names = adapter_names or ["openpyxl", "wolfxl"]
    modules: dict[str, Any] = {"openpyxl": _openpyxl_module()}
    for name in names:
        if name == "openpyxl":
            continue
        if name == "wolfxl":
            try:
                import wolfxl
            except Exception:
                continue
            modules["wolfxl"] = wolfxl
    return modules


def _openpyxl_module() -> Any:
    import openpyxl

    return openpyxl


def _new_workbook(module: Any) -> Any:
    workbook = module.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    return workbook


def _save(workbook: Any, path: Path) -> None:
    workbook.save(path)
    close = getattr(workbook, "close", None)
    if close is not None:
        close()


def _case_create_string(module: Any, path: Path) -> JSONDict:
    wb = _new_workbook(module)
    wb["Sheet1"]["A1"] = "Revenue"
    _save(wb, path)
    return {"A1": "Revenue"}


def _case_create_number(module: Any, path: Path) -> JSONDict:
    wb = _new_workbook(module)
    wb["Sheet1"]["A1"] = 123
    _save(wb, path)
    return {"A1": 123}


def _case_create_formula(module: Any, path: Path) -> JSONDict:
    wb = _new_workbook(module)
    ws = wb["Sheet1"]
    ws["A1"] = 1
    ws["A2"] = 2
    ws["A3"] = "=SUM(A1:A2)"
    _save(wb, path)
    return {"A3": "=SUM(A1:A2)"}


def _case_font_bold(module: Any, path: Path) -> JSONDict:
    from openpyxl.styles import Font

    wb = _new_workbook(module)
    wb["Sheet1"]["A1"] = "Bold"
    wb["Sheet1"]["A1"].font = Font(bold=True)
    _save(wb, path)
    return {"bold": True}


def _case_font_color(module: Any, path: Path) -> JSONDict:
    from openpyxl.styles import Font

    wb = _new_workbook(module)
    wb["Sheet1"]["A1"] = "Red"
    wb["Sheet1"]["A1"].font = Font(color="FF0000")
    _save(wb, path)
    return {"font_color": "FF0000"}


def _case_fill_color(module: Any, path: Path) -> JSONDict:
    from openpyxl.styles import PatternFill

    wb = _new_workbook(module)
    wb["Sheet1"]["A1"] = "Fill"
    wb["Sheet1"]["A1"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    _save(wb, path)
    return {"fill": "FFFF00"}


def _case_number_format(module: Any, path: Path) -> JSONDict:
    wb = _new_workbook(module)
    wb["Sheet1"]["A1"] = 12.34
    wb["Sheet1"]["A1"].number_format = "$#,##0.00"
    _save(wb, path)
    return {"number_format": "$#,##0.00"}


def _case_alignment_wrap(module: Any, path: Path) -> JSONDict:
    from openpyxl.styles import Alignment

    wb = _new_workbook(module)
    wb["Sheet1"]["A1"] = "Wrapped text"
    wb["Sheet1"]["A1"].alignment = Alignment(wrap_text=True)
    _save(wb, path)
    return {"wrap_text": True}


def _case_comment(module: Any, path: Path) -> JSONDict:
    from openpyxl.comments import Comment

    wb = _new_workbook(module)
    wb["Sheet1"]["A1"] = "Reviewed"
    wb["Sheet1"]["A1"].comment = Comment("Tie to PBC", "ExcelBench")
    _save(wb, path)
    return {"comment": "Tie to PBC"}


def _case_hyperlink(module: Any, path: Path) -> JSONDict:
    wb = _new_workbook(module)
    wb["Sheet1"]["A1"] = "Project"
    wb["Sheet1"]["A1"].hyperlink = "https://github.com/SynthGL/wolfxl"
    _save(wb, path)
    return {"hyperlink": "https://github.com/SynthGL/wolfxl"}


def _case_validation(module: Any, path: Path) -> JSONDict:
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = _new_workbook(module)
    ws = wb["Sheet1"]
    ws["A1"] = "Open"
    dv = DataValidation(type="list", formula1='"Open,Closed"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(ws["A1"])
    _save(wb, path)
    return {"validation": "list"}


def _case_merge(module: Any, path: Path) -> JSONDict:
    wb = _new_workbook(module)
    ws = wb["Sheet1"]
    ws["A1"] = "Header"
    ws.merge_cells("A1:C1")
    _save(wb, path)
    return {"merge": "A1:C1"}


def _case_unmerge(module: Any, path: Path) -> JSONDict:
    wb = _new_workbook(module)
    ws = wb["Sheet1"]
    ws.merge_cells("A1:C1")
    ws.unmerge_cells("A1:C1")
    _save(wb, path)
    return {"merge": None}


def _case_freeze(module: Any, path: Path) -> JSONDict:
    wb = _new_workbook(module)
    wb["Sheet1"].freeze_panes = "B2"
    _save(wb, path)
    return {"freeze_panes": "B2"}


def _case_named_range(module: Any, path: Path) -> JSONDict:
    from openpyxl.workbook.defined_name import DefinedName

    wb = _new_workbook(module)
    wb["Sheet1"]["A1"] = "Named"
    wb.defined_names.add(DefinedName("MetricName", attr_text="'Sheet1'!$A$1"))
    _save(wb, path)
    return {"defined_name": "MetricName"}


def _case_table(module: Any, path: Path) -> JSONDict:
    from openpyxl.worksheet.table import Table, TableStyleInfo

    wb = _new_workbook(module)
    ws = wb["Sheet1"]
    ws.append(["Metric", "Value"])
    ws.append(["Revenue", 100])
    table = Table(displayName="CompatTable", ref="A1:B2")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)
    _save(wb, path)
    return {"table": "CompatTable"}


def _case_sheet_create(module: Any, path: Path) -> JSONDict:
    wb = _new_workbook(module)
    wb.create_sheet("Data")
    _save(wb, path)
    return {"sheets": ["Sheet1", "Data"]}


def _case_sheet_rename(module: Any, path: Path) -> JSONDict:
    wb = _new_workbook(module)
    wb["Sheet1"].title = "Renamed"
    _save(wb, path)
    return {"sheet": "Renamed"}


def _case_row_height(module: Any, path: Path) -> JSONDict:
    wb = _new_workbook(module)
    wb["Sheet1"].row_dimensions[1].height = 30
    _save(wb, path)
    return {"row_height": 30}


def _case_column_width(module: Any, path: Path) -> JSONDict:
    wb = _new_workbook(module)
    wb["Sheet1"].column_dimensions["A"].width = 24
    _save(wb, path)
    return {"column_width": 24}


def _stable_json(value: Any) -> str:
    return json.dumps(value, sort_keys=True, default=str)


def _render_compatibility_markdown(payload: JSONDict) -> str:
    lines = [
        "# ExcelBench Compatibility Context",
        "",
        f"- Generated: `{payload['generated_at']}`",
        f"- Passed: `{payload['passed']}`",
        f"- Cases: `{len(payload['cases'])}`",
        "",
        "| Adapter | Case | Surface | Status | Return | Snapshot | Deltas |",
        "|---------|------|---------|--------|--------|----------|--------|",
    ]
    for row in payload["results"]:
        if row["skipped"]:
            status = "skipped"
        elif row["passed"]:
            status = "passed"
        else:
            status = "failed"
        lines.append(
            f"| {row['adapter']} | {row['case_id']} | {row['api_surface']} | {status} | "
            f"{row['return_match']} | {row['snapshot_match']} | {row['delta_count']} |"
        )
    lines.append("")
    return "\n".join(lines)


def _render_compatibility_context(requested: list[str]) -> str:
    return "\n".join(
        [
            "# Compatibility Context",
            "",
            "This lane runs openpyxl-style snippets against compatible adapters.",
            "",
            f"- Requested adapters: {', '.join(requested)}",
            "- openpyxl is the reference implementation.",
            "- Non openpyxl-compatible adapters are explicit skips.",
            "",
        ]
    )
