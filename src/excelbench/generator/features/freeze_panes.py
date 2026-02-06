"""Generator for freeze panes / split views test cases (Tier 2)."""

import sys
from pathlib import Path

import xlwings as xw

from excelbench.generator.base import FeatureGenerator
from excelbench.models import FreezePaneSpec, Importance, TestCase


class FreezePanesGenerator(FeatureGenerator):
    """Generates test cases for freeze panes and split views."""

    feature_name = "freeze_panes"
    tier = 2
    filename = "17_freeze_panes.xlsx"

    def __init__(self) -> None:
        self._use_openpyxl = sys.platform == "darwin"
        self._ops: list[dict[str, object]] = []

    def generate(self, sheet: xw.Sheet) -> list[TestCase]:
        self.setup_header(sheet)

        is_mac = sys.platform == "darwin"

        test_cases: list[TestCase] = []
        row = 2

        wb = sheet.book
        window = wb.app.api.active_window if not is_mac else None

        def set_window_attr(prop: str, value: object) -> None:
            if window is None:
                return
            setattr(window, prop, value)

        freeze_b2 = wb.sheets.add("FreezeB2")
        freeze_d5 = wb.sheets.add("FreezeD5")
        split_sheet = wb.sheets.add("SplitPanes")

        # Freeze row+col at B2
        label = "Freeze panes at B2"
        if is_mac:
            self._ops.append({"sheet": "FreezeB2", "mode": "freeze", "top_left_cell": "B2"})
        else:
            freeze_b2.activate()
            freeze_b2.range("B2").select()
            set_window_attr("FreezePanes", True)
        expected = FreezePaneSpec(
            mode="freeze",
            top_left_cell="B2",
        ).to_expected()
        self.write_test_case(sheet, row, label, expected)
        test_cases.append(
            TestCase(
                id="freeze_b2",
                label=label,
                row=row,
                expected=expected,
                sheet="FreezeB2",
            )
        )
        row += 1

        # Freeze at D5 (non-A1 top-left)
        label = "Freeze panes at D5"
        if is_mac:
            self._ops.append({"sheet": "FreezeD5", "mode": "freeze", "top_left_cell": "D5"})
        else:
            set_window_attr("FreezePanes", False)
            freeze_d5.activate()
            freeze_d5.range("D5").select()
            set_window_attr("FreezePanes", True)
        expected = FreezePaneSpec(
            mode="freeze",
            top_left_cell="D5",
        ).to_expected()
        self.write_test_case(sheet, row, label, expected)
        test_cases.append(
            TestCase(
                id="freeze_d5",
                label=label,
                row=row,
                expected=expected,
                sheet="FreezeD5",
                importance=Importance.EDGE,
            )
        )
        row += 1

        # Split panes (not freeze)
        label = "Split panes row=2 col=1"
        if is_mac:
            self._ops.append({"sheet": "SplitPanes", "mode": "split", "x_split": 1, "y_split": 2})
        else:
            set_window_attr("FreezePanes", False)
            split_sheet.activate()
            set_window_attr("SplitRow", 2)
            set_window_attr("SplitColumn", 1)
        expected = FreezePaneSpec(
            mode="split",
            x_split=1,
            y_split=2,
        ).to_expected()
        self.write_test_case(sheet, row, label, expected)
        test_cases.append(
            TestCase(
                id="split_2x1",
                label=label,
                row=row,
                expected=expected,
                sheet="SplitPanes",
                importance=Importance.EDGE,
            )
        )

        return test_cases

    def post_process(self, output_path: Path) -> None:
        if not self._use_openpyxl or not self._ops:
            return
        from openpyxl import load_workbook
        from openpyxl.worksheet.views import Pane

        wb = load_workbook(output_path)
        for op in self._ops:
            sheet_name = op.get("sheet")
            if not isinstance(sheet_name, str) or sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            mode = op.get("mode")
            if mode == "freeze":
                ws.freeze_panes = op.get("top_left_cell")
            elif mode == "split":
                ws.freeze_panes = None
                pane = ws.sheet_view.pane
                if pane is None:
                    pane = Pane()
                    ws.sheet_view.pane = pane
                pane.xSplit = op.get("x_split")
                pane.ySplit = op.get("y_split")
                pane.topLeftCell = op.get("top_left_cell")
                pane.state = "split"
        wb.save(output_path)
