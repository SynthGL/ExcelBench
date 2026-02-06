"""Generator for alignment test cases."""

import sys
from pathlib import Path

import xlwings as xw

from excelbench.generator.base import FeatureGenerator
from excelbench.models import TestCase


class XlHAlign:
    LEFT = -4131
    CENTER = -4108
    RIGHT = -4152


class XlVAlign:
    TOP = -4160
    CENTER = -4108
    BOTTOM = -4107


class AlignmentGenerator(FeatureGenerator):
    """Generates test cases for cell alignment."""

    feature_name = "alignment"
    tier = 1
    filename = "06_alignment.xlsx"

    def __init__(self) -> None:
        self._use_openpyxl = sys.platform == "darwin"
        self._ops: list[dict[str, object]] = []

    def generate(self, sheet: xw.Sheet) -> list[TestCase]:
        self.setup_header(sheet)

        test_cases: list[TestCase] = []
        row = 2

        test_cases.append(self._test_h_align(sheet, row, "Align - left", XlHAlign.LEFT, "left"))
        row += 1
        test_cases.append(
            self._test_h_align(sheet, row, "Align - center", XlHAlign.CENTER, "center")
        )
        row += 1
        test_cases.append(self._test_h_align(sheet, row, "Align - right", XlHAlign.RIGHT, "right"))
        row += 1

        test_cases.append(self._test_v_align(sheet, row, "Align - top", XlVAlign.TOP, "top"))
        row += 1
        test_cases.append(
            self._test_v_align(sheet, row, "Align - center", XlVAlign.CENTER, "center")
        )
        row += 1
        test_cases.append(
            self._test_v_align(sheet, row, "Align - bottom", XlVAlign.BOTTOM, "bottom")
        )
        row += 1

        test_cases.append(self._test_wrap(sheet, row))
        row += 1
        test_cases.append(self._test_rotation(sheet, row))
        row += 1
        test_cases.append(self._test_indent(sheet, row))
        row += 1

        return test_cases

    def _test_h_align(
        self,
        sheet: xw.Sheet,
        row: int,
        label: str,
        value: int,
        expected: str,
    ) -> TestCase:
        exp = {"h_align": expected}
        self.write_test_case(sheet, row, label, exp)
        cell = sheet.range(f"B{row}")
        cell.value = label
        if self._use_openpyxl:
            self._ops.append({"cell": f"B{row}", "horizontal": expected})
        else:
            self._set_horizontal_alignment(cell, value)
        return TestCase(id=f"h_{expected}", label=label, row=row, expected=exp)

    def _test_v_align(
        self,
        sheet: xw.Sheet,
        row: int,
        label: str,
        value: int,
        expected: str,
    ) -> TestCase:
        exp = {"v_align": expected}
        self.write_test_case(sheet, row, label, exp)
        cell = sheet.range(f"B{row}")
        cell.value = label
        if self._use_openpyxl:
            self._ops.append({"cell": f"B{row}", "vertical": expected})
        else:
            self._set_vertical_alignment(cell, value)
        return TestCase(id=f"v_{expected}", label=label, row=row, expected=exp)

    def _test_wrap(self, sheet: xw.Sheet, row: int) -> TestCase:
        label = "Align - wrap text"
        exp = {"wrap": True}
        self.write_test_case(sheet, row, label, exp)
        cell = sheet.range(f"B{row}")
        cell.value = "Line 1\nLine 2"
        if self._use_openpyxl:
            self._ops.append({"cell": f"B{row}", "wrapText": True})
        else:
            self._set_wrap(cell, True)
        return TestCase(id="wrap_text", label=label, row=row, expected=exp)

    def _test_rotation(self, sheet: xw.Sheet, row: int) -> TestCase:
        label = "Align - rotation 45"
        exp = {"rotation": 45}
        self.write_test_case(sheet, row, label, exp)
        cell = sheet.range(f"B{row}")
        cell.value = "Rotated"
        if self._use_openpyxl:
            self._ops.append({"cell": f"B{row}", "textRotation": 45})
        else:
            self._set_orientation(cell, 45)
        return TestCase(id="rotation_45", label=label, row=row, expected=exp)

    def _test_indent(self, sheet: xw.Sheet, row: int) -> TestCase:
        label = "Align - indent 2"
        exp = {"indent": 2}
        self.write_test_case(sheet, row, label, exp)
        cell = sheet.range(f"B{row}")
        cell.value = "Indented"
        if self._use_openpyxl:
            self._ops.append({"cell": f"B{row}", "indent": 2})
        else:
            self._set_indent(cell, 2)
        return TestCase(id="indent_2", label=label, row=row, expected=exp)

    def post_process(self, output_path: Path) -> None:
        if not self._use_openpyxl or not self._ops:
            return
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment

        wb = load_workbook(output_path)
        ws = wb[self.feature_name]
        for op in self._ops:
            cell_ref = op.pop("cell")
            ws[cell_ref].alignment = Alignment(**op)
        wb.save(output_path)

    @staticmethod
    def _set_horizontal_alignment(cell: xw.Range, value: int) -> None:
        try:
            cell.api.HorizontalAlignment = value
        except Exception:
            cell.api.horizontal_alignment.set(value)

    @staticmethod
    def _set_vertical_alignment(cell: xw.Range, value: int) -> None:
        try:
            cell.api.VerticalAlignment = value
        except Exception:
            cell.api.vertical_alignment.set(value)

    @staticmethod
    def _set_wrap(cell: xw.Range, value: bool) -> None:
        try:
            cell.api.WrapText = value
        except Exception:
            cell.api.wrap_text.set(value)

    @staticmethod
    def _set_orientation(cell: xw.Range, value: int) -> None:
        try:
            cell.api.Orientation = value
        except Exception:
            cell.api.text_orientation.set(value)

    @staticmethod
    def _set_indent(cell: xw.Range, value: int) -> None:
        try:
            cell.api.IndentLevel = value
        except Exception:
            cell.api.indent_level.set(value)
