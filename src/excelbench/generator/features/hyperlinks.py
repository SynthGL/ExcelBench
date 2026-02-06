"""Generator for hyperlink test cases (Tier 2)."""

import sys
from pathlib import Path

import xlwings as xw

from excelbench.generator.base import FeatureGenerator
from excelbench.models import HyperlinkSpec, Importance, TestCase


class HyperlinksGenerator(FeatureGenerator):
    """Generates test cases for hyperlinks."""

    feature_name = "hyperlinks"
    tier = 2
    filename = "13_hyperlinks.xlsx"

    def __init__(self) -> None:
        self._use_openpyxl = sys.platform == "darwin"
        self._ops: list[dict[str, object]] = []

    def generate(self, sheet: xw.Sheet) -> list[TestCase]:
        self.setup_header(sheet)

        test_cases: list[TestCase] = []
        row = 2

        # External hyperlink
        label = "Hyperlink: external URL"
        cell = "B2"
        if not self._use_openpyxl:
            sheet.range(cell).add_hyperlink(
                "https://example.com/docs",
                "Example Docs",
                "Go to docs",
            )
        self._ops.append({
            "cell": cell, "target": "https://example.com/docs",
            "display": "Example Docs", "tooltip": "Go to docs",
        })
        expected = HyperlinkSpec(
            cell=cell,
            target="https://example.com/docs",
            display="Example Docs",
            tooltip="Go to docs",
            internal=False,
        ).to_expected()
        self.write_test_case(sheet, row, label, expected)
        test_cases.append(TestCase(id="link_external", label=label, row=row, expected=expected))
        row += 1

        # Internal hyperlink to another sheet
        label = "Hyperlink: internal sheet"
        target_sheet = sheet.book.sheets.add("Targets")
        target_sheet.range("A1").value = "Target"
        cell = "B3"
        if not self._use_openpyxl:
            sheet.range(cell).add_hyperlink(
                "#'Targets'!A1",
                "Go Target",
                "Jump to target",
            )
        self._ops.append({
            "cell": cell, "location": "'Targets'!A1",
            "display": "Go Target", "tooltip": "Jump to target",
        })
        expected = HyperlinkSpec(
            cell=cell,
            target="Targets!A1",
            display="Go Target",
            tooltip="Jump to target",
            internal=True,
        ).to_expected()
        self.write_test_case(sheet, row, label, expected)
        test_cases.append(TestCase(
            id="link_internal",
            label=label,
            row=row,
            expected=expected,
            importance=Importance.EDGE,
        ))
        row += 1

        # Mailto hyperlink
        label = "Hyperlink: mailto"
        cell = "B4"
        if not self._use_openpyxl:
            sheet.range(cell).add_hyperlink(
                "mailto:test@example.com",
                "Email",
                "Send email",
            )
        self._ops.append({
            "cell": cell, "target": "mailto:test@example.com",
            "display": "Email", "tooltip": "Send email",
        })
        expected = HyperlinkSpec(
            cell=cell,
            target="mailto:test@example.com",
            display="Email",
            tooltip="Send email",
            internal=False,
        ).to_expected()
        self.write_test_case(sheet, row, label, expected)
        test_cases.append(TestCase(id="link_mailto", label=label, row=row, expected=expected))
        row += 1

        # Long URL with encoding
        label = "Hyperlink: long encoded URL"
        cell = "B5"
        url = "https://example.com/search?q=excel%20bench&sort=desc#section-2"
        if not self._use_openpyxl:
            sheet.range(cell).add_hyperlink(url, "Search", "Encoded URL")
        self._ops.append({
            "cell": cell, "target": url,
            "display": "Search", "tooltip": "Encoded URL",
        })
        expected = HyperlinkSpec(
            cell=cell,
            target=url,
            display="Search",
            tooltip="Encoded URL",
            internal=False,
        ).to_expected()
        self.write_test_case(sheet, row, label, expected)
        test_cases.append(TestCase(
            id="link_long",
            label=label,
            row=row,
            expected=expected,
            importance=Importance.EDGE,
        ))

        return test_cases

    def post_process(self, output_path: Path) -> None:
        if not self._use_openpyxl or not self._ops:
            return
        from openpyxl import load_workbook
        from openpyxl.worksheet.hyperlink import Hyperlink

        wb = load_workbook(output_path)
        ws = wb[self.feature_name]
        for op in self._ops:
            cell = str(op["cell"])
            display = op.get("display")
            if display:
                ws[cell].value = display
            target = op.get("target")
            location = op.get("location")
            if location:
                ws[cell].hyperlink = Hyperlink(ref=cell, location=str(location))
            elif target:
                ws[cell].hyperlink = str(target)
            tooltip = op.get("tooltip")
            if tooltip and ws[cell].hyperlink:
                ws[cell].hyperlink.tooltip = str(tooltip)
        wb.save(output_path)
