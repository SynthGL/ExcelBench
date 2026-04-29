from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill

from excelbench.harness.semantic_diff import diff_workbooks, write_diff_artifacts
from excelbench.harness.workbook_snapshot import snapshot_workbook


def _write_workbook(path: Path, *, value: str = "Revenue", fill: str = "FFFF00") -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet["A1"] = value
    sheet["A1"].fill = PatternFill(fill_type="solid", fgColor=fill)
    sheet["B1"] = "Link"
    sheet["B1"].hyperlink = "https://example.com"
    sheet["C1"].comment = Comment("reviewed", "ExcelBench")
    sheet.merge_cells("D1:E1")
    sheet.freeze_panes = "B2"
    workbook.save(path)
    workbook.close()


def test_snapshot_workbook_captures_semantic_categories(tmp_path: Path) -> None:
    path = tmp_path / "book.xlsx"
    _write_workbook(path)

    snapshot = snapshot_workbook(path)

    assert "cells" in snapshot.categories
    assert snapshot.categories["cells"]["Data"]["A1"]["value"] == "Revenue"
    assert snapshot.categories["merges"]["Data"] == ["D1:E1"]
    assert snapshot.categories["hyperlinks"]["Data"]["B1"]["target"] == "https://example.com"
    assert snapshot.categories["comments"]["Data"]["C1"]["text"] == "reviewed"
    assert snapshot.categories["freeze_panes"]["Data"] == "B2"


def test_diff_workbooks_reports_category_counts(tmp_path: Path) -> None:
    left = tmp_path / "left.xlsx"
    right = tmp_path / "right.xlsx"
    _write_workbook(left, value="Revenue")
    _write_workbook(right, value="COGS")

    diff = diff_workbooks(left, right)

    assert not diff.passed
    assert diff.category_counts()["cells"] >= 1


def test_write_diff_artifacts(tmp_path: Path) -> None:
    left = tmp_path / "left.xlsx"
    right = tmp_path / "right.xlsx"
    output = tmp_path / "diff"
    _write_workbook(left, value="Revenue")
    _write_workbook(right, value="COGS")

    write_diff_artifacts(left, right, output)

    summary = json.loads((output / "summary.json").read_text())
    assert summary["passed"] is False
    assert (output / "summary.md").exists()
    assert (output / "categories" / "cells.json").exists()
