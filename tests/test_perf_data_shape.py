"""Tests for the Sprint 2 data-shape extension.

Covers the new ``value_type`` branches in ``_run_workload_write`` (float, date,
datetime, boolean, formula_simple, formula_cross_sheet, mixed_realistic) plus
the CLI helpers ``_resolve_shape_features`` / ``_shape_fixtures_stale``.

Each value_type test runs one ``run_perf`` iteration with a 4-cell write
workload against ``OpenpyxlAdapter``. The asserted shape is intentionally
loose — we only check the workload completed and op_count was recorded — so
the tests stay portable across adapter quirks (some libs auto-evaluate
formulas, others store the string verbatim).
"""

from __future__ import annotations

import json
from datetime import UTC, datetime
from pathlib import Path

import pytest
from openpyxl import Workbook

from excelbench.generator.generate import write_manifest
from excelbench.harness.adapters.openpyxl_adapter import OpenpyxlAdapter
from excelbench.models import Importance, Manifest
from excelbench.models import TestCase as BenchCase
from excelbench.models import TestFile as BenchFile
from excelbench.perf.runner import run_perf


def _build_shape_suite(
    tmp_path: Path,
    scenario: str,
    *,
    value_type: str,
    extra: dict[str, object] | None = None,
) -> Path:
    """Create a manifest with one bulk_write_grid workload for the given value_type.

    The companion .xlsx is a tiny 2x2 sheet, so the read iteration succeeds
    even though we only care about the write side.
    """
    suite = tmp_path / "suite"
    suite.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "S1"
    # Pre-populate Sheet2 so formula_cross_sheet has a valid target on read-back.
    ws2 = wb.create_sheet("Sheet2")
    for r in range(1, 3):
        for c in range(1, 3):
            ws.cell(row=r, column=c, value=r * 10 + c)
            ws2.cell(row=r, column=c, value=99)

    (suite / "tier0").mkdir(parents=True, exist_ok=True)
    wb_path = suite / "tier0" / f"00_{scenario}.xlsx"
    wb.save(wb_path)

    workload: dict[str, object] = {
        "scenario": scenario,
        "op": "bulk_write_grid",
        "sheet": "S1",
        "range": "A1:B2",
        "value_type": value_type,
        "start": 1,
        "step": 1,
    }
    if extra:
        workload.update(extra)

    manifest = Manifest(
        generated_at=datetime.now(UTC),
        excel_version="test",
        generator_version="test",
        file_format="xlsx",
        files=[
            BenchFile(
                path=f"tier0/00_{scenario}.xlsx",
                feature=scenario,
                tier=0,
                file_format="xlsx",
                test_cases=[
                    BenchCase(
                        id=scenario,
                        label=f"Throughput: {scenario}",
                        row=1,
                        expected={"workload": workload},
                        importance=Importance.BASIC,
                    )
                ],
            )
        ],
    )
    write_manifest(manifest, suite / "manifest.json")
    return suite


@pytest.mark.parametrize(
    ("value_type", "extra"),
    [
        ("number", None),
        ("float", None),
        ("string", {"string_length": 16}),
        ("string", {"string_length": 256, "string_mode": "repeated"}),
        ("date", None),
        ("datetime", None),
        ("boolean", None),
        ("formula_simple", None),
        ("formula_cross_sheet", None),
        ("mixed_realistic", None),
    ],
)
def test_data_shape_value_type_writes(
    tmp_path: Path, value_type: str, extra: dict[str, object] | None
) -> None:
    scenario = f"data_shape_{value_type}_{(extra or {}).get('string_length', 'd')}"
    suite = _build_shape_suite(tmp_path, scenario, value_type=value_type, extra=extra)

    results = run_perf(
        suite,
        adapters=[OpenpyxlAdapter()],
        warmup=0,
        iters=1,
        breakdown=False,
    )

    row = results.results[0]
    assert row.feature == scenario
    assert row.perf["write"] is not None
    assert row.perf["write"].op_count == 4
    assert row.perf["write"].op_unit == "cells"


def test_data_shape_unsupported_value_type_surfaces_in_notes(tmp_path: Path) -> None:
    """The catch-all branch raises ValueError; ``run_perf`` records it in notes."""
    scenario = "data_shape_garbage"
    suite = _build_shape_suite(tmp_path, scenario, value_type="bogus_type_xyz")

    results = run_perf(
        suite,
        adapters=[OpenpyxlAdapter()],
        warmup=0,
        iters=1,
        breakdown=False,
    )
    row = results.results[0]
    assert row.notes is not None
    assert "bogus_type_xyz" in row.notes or "value_type" in row.notes.lower()


def test_data_shape_sparse_skips_cells(tmp_path: Path) -> None:
    """sparse_every shrinks the recorded op_count to filled cells only."""
    scenario = "data_shape_sparse"
    suite = _build_shape_suite(
        tmp_path,
        scenario,
        value_type="number",
        extra={"sparse_every": 2, "range": "A1:D1"},
    )

    results = run_perf(
        suite,
        adapters=[OpenpyxlAdapter()],
        warmup=0,
        iters=1,
        breakdown=False,
    )

    row = results.results[0]
    # 4 cells with sparse_every=2 => 2 filled cells.
    write_op = row.perf["write"]
    assert write_op is not None
    assert write_op.op_count == 2


# ---------------------------------------------------------------------------
# CLI helpers: _resolve_shape_features and _shape_fixtures_stale.
# ---------------------------------------------------------------------------


def test_resolve_shape_features_all_dtypes_at_100k() -> None:
    from excelbench.cli import _DATA_SHAPE_DTYPES, _resolve_shape_features

    reads, writes, tiers = _resolve_shape_features(types_arg="all", rows=100_000)
    # 10 dtypes × 3 tiers (1k/10k/100k) = 30 read + 30 write features.
    assert len(reads) == 10 * 3
    assert len(writes) == 10 * 3
    assert tiers == ["1k", "10k", "100k"]
    # Sanity: every dtype shows up at least once.
    for dtype in _DATA_SHAPE_DTYPES:
        assert any(f"_{dtype}_" in r for r in reads), f"dtype {dtype} missing from reads"


def test_resolve_shape_features_filtered_dtype_subset() -> None:
    from excelbench.cli import _resolve_shape_features

    reads, writes, tiers = _resolve_shape_features(
        types_arg="int,formula_simple", rows=10_000
    )
    # 2 dtypes × 2 tiers (1k+10k) = 4 read + 4 write features.
    assert len(reads) == 4
    assert len(writes) == 4
    assert tiers == ["1k", "10k"]
    assert all("_int_" in r or "_formula_simple_" in r for r in reads)


def test_resolve_shape_features_unknown_dtype_raises() -> None:
    from excelbench.cli import _resolve_shape_features

    with pytest.raises(ValueError, match="Unknown --types"):
        _resolve_shape_features(types_arg="int,not_a_real_dtype", rows=10_000)


def test_resolve_shape_features_rows_below_smallest_tier_raises() -> None:
    from excelbench.cli import _resolve_shape_features

    with pytest.raises(ValueError, match="below the smallest tier"):
        _resolve_shape_features(types_arg="all", rows=100)


def test_resolve_shape_features_includes_1m_tier() -> None:
    from excelbench.cli import _resolve_shape_features

    _, _, tiers = _resolve_shape_features(types_arg="int", rows=1_000_000)
    assert "1m" in tiers


def test_shape_fixtures_stale_missing_manifest(tmp_path: Path) -> None:
    from excelbench.cli import _shape_fixtures_stale

    manifest = tmp_path / "manifest.json"
    generator = tmp_path / "gen.py"
    generator.write_text("# generator\n")

    assert _shape_fixtures_stale(manifest, generator, needs_1m=False) is True


def test_shape_fixtures_stale_generator_newer_than_manifest(tmp_path: Path) -> None:
    from excelbench.cli import _shape_fixtures_stale

    manifest = tmp_path / "manifest.json"
    generator = tmp_path / "gen.py"

    # Create manifest first (older), then generator (newer).
    manifest.write_text('{"files": []}')
    import os
    import time

    time.sleep(0.01)
    generator.write_text("# generator\n")
    # Force generator's mtime to be strictly later.
    later = manifest.stat().st_mtime + 5.0
    os.utime(generator, (later, later))

    assert _shape_fixtures_stale(manifest, generator, needs_1m=False) is True


def test_shape_fixtures_stale_needs_1m_but_missing(tmp_path: Path) -> None:
    from excelbench.cli import _shape_fixtures_stale

    manifest = tmp_path / "manifest.json"
    generator = tmp_path / "gen.py"
    generator.write_text("# generator\n")
    # Manifest with only 1k tier — needs_1m should trigger regeneration.
    manifest.write_text(
        json.dumps(
            {
                "files": [
                    {"feature": "data_shape_int_1k_bulk_read", "path": "x.xlsx"}
                ]
            }
        )
    )
    # Bump manifest mtime above generator so the mtime check doesn't trigger.
    import os

    later = generator.stat().st_mtime + 5.0
    os.utime(manifest, (later, later))

    assert _shape_fixtures_stale(manifest, generator, needs_1m=True) is True


def test_shape_fixtures_stale_fresh_with_1m(tmp_path: Path) -> None:
    from excelbench.cli import _shape_fixtures_stale

    manifest = tmp_path / "manifest.json"
    generator = tmp_path / "gen.py"
    generator.write_text("# generator\n")
    manifest.write_text(
        json.dumps(
            {
                "files": [
                    {"feature": "data_shape_int_1m_bulk_read", "path": "x.xlsx"}
                ]
            }
        )
    )
    # Bump manifest mtime above generator.
    import os

    later = generator.stat().st_mtime + 5.0
    os.utime(manifest, (later, later))

    assert _shape_fixtures_stale(manifest, generator, needs_1m=True) is False
    assert _shape_fixtures_stale(manifest, generator, needs_1m=False) is False


def test_shape_fixtures_stale_file_shape_only_manifest_is_stale(tmp_path: Path) -> None:
    """Cross-command guard: a manifest written by perf-file-shape (file_shape
    only) must be treated as stale by perf-shape, regardless of needs_1m or
    fresh mtime, so the data-shape fixtures get regenerated.
    """
    import os

    from excelbench.cli import _shape_fixtures_stale

    manifest = tmp_path / "manifest.json"
    manifest.write_text(
        json.dumps(
            {
                "files": [
                    {"feature": "file_shape_wide_10k_bulk_read", "path": "x.xlsx"},
                    {"feature": "file_shape_wide_10k_bulk_write", "path": "x.xlsx"},
                ]
            }
        )
    )
    generator = tmp_path / "gen.py"
    generator.write_text("# generator")
    later = generator.stat().st_mtime + 5.0
    os.utime(manifest, (later, later))
    # Both flag values must report stale: the manifest contains zero
    # data_shape_* entries even though it's fresh on disk.
    assert _shape_fixtures_stale(manifest, generator, needs_1m=False) is True
    assert _shape_fixtures_stale(manifest, generator, needs_1m=True) is True


# ---------------------------------------------------------------------------
# Dashboard helper: _section_data_shape (renders read/write heatmaps).
# ---------------------------------------------------------------------------


def _shape_perf_payload(
    *, libs: list[str], tiers: list[str], dtypes: list[str]
) -> dict[str, object]:
    """Synthesize a perf dict shaped like results.json for dashboard tests."""
    results = []
    for lib in libs:
        for dtype in dtypes:
            for tier in tiers:
                for op in ("read", "write"):
                    feature = f"data_shape_{dtype}_{tier}_bulk_{op}"
                    # Make wolfxl artificially fast so library ordering is stable.
                    base_ms = 1.0 if lib == "wolfxl" else 5.0
                    results.append(
                        {
                            "feature": feature,
                            "library": lib,
                            "perf": {
                                op: {
                                    "wall_ms": {"min": base_ms, "p50": base_ms, "p95": base_ms},
                                }
                            },
                        }
                    )
    return {"results": results}


# ---------------------------------------------------------------------------
# CLI: perf_shape command direct invocation.
# ---------------------------------------------------------------------------


def _build_shape_fixture_manifest(fixtures_dir: Path, dtype: str, tier: str) -> None:
    """Synthesize a minimal data_shape_* manifest the perf_shape CLI can consume.

    Skips the expensive xlsxwriter generation by writing one fake bulk_read +
    one bulk_write feature with a tiny range. The CLI sees this as a fresh
    manifest and won't shell out to the generator (mtime is newer than gen).
    """
    fixtures_dir.mkdir(parents=True, exist_ok=True)
    tier_dir = fixtures_dir / "data_shape"
    tier_dir.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "S1"
    ws["A1"] = 1
    ws["B1"] = 2
    wb_path = tier_dir / f"data_shape_{dtype}_{tier}.xlsx"
    wb.save(wb_path)

    files = []
    for op in ("read", "write"):
        feature = f"data_shape_{dtype}_{tier}_bulk_{op}"
        files.append(
            BenchFile(
                path=f"data_shape/data_shape_{dtype}_{tier}.xlsx",
                feature=feature,
                tier=0,
                file_format="xlsx",
                test_cases=[
                    BenchCase(
                        id=feature,
                        label=feature,
                        row=1,
                        expected={
                            "workload": {
                                "scenario": feature,
                                "op": "bulk_write_grid" if op == "write" else "bulk_sheet_values",
                                "sheet": "S1",
                                "range": "A1:B1",
                                "value_type": dtype if dtype in {"number", "string"} else "number",
                                "operations": [op],
                            }
                        },
                        importance=Importance.BASIC,
                    )
                ],
            )
        )

    manifest = Manifest(
        generated_at=datetime.now(UTC),
        excel_version="test",
        generator_version="test",
        file_format="xlsx",
        files=files,
    )
    write_manifest(manifest, fixtures_dir / "manifest.json")


def test_perf_shape_command_invalid_memory_mode_exits(tmp_path: Path) -> None:
    import typer

    from excelbench.cli import perf_shape

    with pytest.raises(typer.Exit) as exc:
        perf_shape(
            rows=1_000,
            types="int",
            output=tmp_path / "out",
            fixtures=tmp_path / "fixtures",
            adapters=None,
            warmup=0,
            iters=1,
            iteration_policy="fixed",
            breakdown=False,
            memory_mode="not_a_real_mode",
            regenerate=False,
        )
    assert exc.value.exit_code == 1


def test_perf_shape_command_invalid_dtype_exits(tmp_path: Path) -> None:
    import typer

    from excelbench.cli import perf_shape

    # Use existing manifest so we don't trigger generator subprocess.
    _build_shape_fixture_manifest(tmp_path / "fixtures", dtype="number", tier="1k")

    with pytest.raises(typer.Exit) as exc:
        perf_shape(
            rows=1_000,
            types="not_a_real_dtype",
            output=tmp_path / "out",
            fixtures=tmp_path / "fixtures",
            adapters=None,
            warmup=0,
            iters=1,
            iteration_policy="fixed",
            breakdown=False,
            memory_mode="getrusage",
            regenerate=False,
        )
    assert exc.value.exit_code == 1


def test_perf_shape_command_happy_path_writes_outputs(tmp_path: Path) -> None:
    """End-to-end: perf_shape walks the manifest, runs run_perf, writes results."""
    from excelbench.cli import perf_shape

    fixtures = tmp_path / "fixtures"
    output = tmp_path / "out"
    _build_shape_fixture_manifest(fixtures, dtype="int", tier="1k")

    perf_shape(
        rows=1_000,
        types="int",  # maps to data_shape_int_1k_bulk_*
        output=output,
        fixtures=fixtures,
        adapters=["openpyxl"],
        warmup=0,
        iters=1,
        iteration_policy="fixed",
        breakdown=False,
        memory_mode="getrusage",
        regenerate=False,
    )

    assert (output / "perf" / "results.json").exists()
    assert (output / "perf" / "matrix.csv").exists()
    assert (output / "perf" / "history.jsonl").exists()


def test_perf_shape_command_unknown_adapter_exits(tmp_path: Path) -> None:
    import typer

    from excelbench.cli import perf_shape

    _build_shape_fixture_manifest(tmp_path / "fixtures", dtype="number", tier="1k")

    with pytest.raises(typer.Exit) as exc:
        perf_shape(
            rows=1_000,
            types="int",
            output=tmp_path / "out",
            fixtures=tmp_path / "fixtures",
            adapters=["nonexistent_lib_xyz"],
            warmup=0,
            iters=1,
            iteration_policy="fixed",
            breakdown=False,
            memory_mode="getrusage",
            regenerate=False,
        )
    assert exc.value.exit_code == 1


def test_section_data_shape_empty_returns_blank() -> None:
    from excelbench.results.html_dashboard import _section_data_shape

    assert _section_data_shape(None) == ""
    assert _section_data_shape({}) == ""
    assert _section_data_shape({"results": []}) == ""


def test_section_data_shape_renders_heatmap_with_real_shape() -> None:
    from excelbench.results.html_dashboard import _section_data_shape

    perf = _shape_perf_payload(
        libs=["wolfxl", "openpyxl"],
        tiers=["1k", "10k"],
        dtypes=["int", "string_short", "formula_simple"],
    )
    html = _section_data_shape(perf)
    assert html, "expected non-empty HTML when results contain data_shape_* features"
    assert 'id="data-shape"' in html
    assert "<h2>Data Shape" in html
    # Both Read and Write subsections render.
    assert ">Read<" in html
    assert ">Write<" in html
    # Library names show up.
    assert "wolfxl" in html
    assert "openpyxl" in html
    # Dtype columns appear.
    for dtype in ("int", "string_short", "formula_simple"):
        assert dtype in html


def test_section_data_shape_skips_unknown_features() -> None:
    """Non-shape feature names should not render the section."""
    from excelbench.results.html_dashboard import _section_data_shape

    perf = {
        "results": [
            {
                "feature": "cell_values",  # not a data_shape_* feature
                "library": "wolfxl",
                "perf": {
                    "read": {"wall_ms": {"min": 1.0, "p50": 1.0, "p95": 1.0}},
                },
            }
        ]
    }
    assert _section_data_shape(perf) == ""


def test_section_data_shape_picks_largest_tier_for_headline() -> None:
    """The dashboard uses the largest tier with data for the headline ms/100k value."""
    from excelbench.results.html_dashboard import _section_data_shape

    perf = _shape_perf_payload(
        libs=["wolfxl"],
        tiers=["1k", "10k", "100k"],
        dtypes=["int"],
    )
    html = _section_data_shape(perf)
    # Tooltip should mention the largest tier (100k) as the headline cell.
    assert "largest tier: 100k" in html


def test_data_shape_color_clamps_min_max_to_endpoints() -> None:
    """Colors interpolate log-scale between green (fast) and red (slow)."""
    from excelbench.results.html_dashboard import _data_shape_color

    fast = _data_shape_color(1.0, col_min=1.0, col_max=100.0)
    slow = _data_shape_color(100.0, col_min=1.0, col_max=100.0)
    # Single-value column (min == max) should not crash.
    flat = _data_shape_color(5.0, col_min=5.0, col_max=5.0)

    for color in (fast, slow, flat):
        assert color.startswith("hsl(") and color.endswith(")")


def test_shape_fixtures_stale_corrupt_manifest_with_needs_1m(tmp_path: Path) -> None:
    from excelbench.cli import _shape_fixtures_stale

    manifest = tmp_path / "manifest.json"
    generator = tmp_path / "gen.py"
    generator.write_text("# generator\n")
    manifest.write_text("not-json{{{")
    # Bump manifest mtime above generator.
    import os

    later = generator.stat().st_mtime + 5.0
    os.utime(manifest, (later, later))

    # Corrupt manifest with needs_1m=True triggers regeneration.
    assert _shape_fixtures_stale(manifest, generator, needs_1m=True) is True


def test_perf_csv_includes_regression_status(tmp_path: Path) -> None:
    from excelbench.perf.renderer import render_perf_csv
    from excelbench.perf.runner import (
        PerfConfig,
        PerfFeatureResult,
        PerfMetadata,
        PerfOpResult,
        PerfResults,
        PerfRunEnvironment,
        PerfStats,
    )

    stats = PerfStats(min=1, p50=1, p95=2)
    res = PerfResults(
        metadata=PerfMetadata(
            benchmark_version="x",
            run_date=datetime.now(UTC),
            excel_version="x",
            platform="x",
            profile="xlsx",
            python="3",
            commit=None,
            config=PerfConfig(warmup=0, iters=1, iteration_policy="fixed", breakdown=False),
            run_environment=PerfRunEnvironment(cpu_model=None, core_count=1, memory_total_mb=None),
        ),
        libraries={"openpyxl": {"capabilities": ["read"]}},
        results=[
            PerfFeatureResult(
                feature="f",
                library="openpyxl",
                workload_size="tiny",
                perf={"read": PerfOpResult(wall_ms=stats, cpu_ms=stats), "write": None},
            )
        ],
    )
    out = tmp_path / "m.csv"
    render_perf_csv(res, out)
    txt = out.read_text()
    assert "regression_status" in txt
    assert "confidence_note" in txt
