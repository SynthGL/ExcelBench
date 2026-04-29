"""Tests for the ApachePoiAdapter cross-language write path."""

from __future__ import annotations

from pathlib import Path

import pytest

from excelbench.harness.adapters.apache_poi_adapter import ApachePoiAdapter
from excelbench.harness.adapters.openpyxl_adapter import OpenpyxlAdapter
from excelbench.models import CellFormat, CellType, CellValue


@pytest.fixture
def opxl() -> OpenpyxlAdapter:
    return OpenpyxlAdapter()


@pytest.fixture
def poi() -> ApachePoiAdapter:
    return ApachePoiAdapter()


class TestApachePoiAdapterInfo:
    def test_name(self, poi: ApachePoiAdapter) -> None:
        assert poi.info.name == "apache-poi"

    def test_capabilities(self, poi: ApachePoiAdapter) -> None:
        assert poi.info.capabilities == {"write"}
        assert poi.can_write()
        assert not poi.can_read()


class TestApachePoiAvailability:
    def test_save_raises_structured_error_when_helper_missing(
        self, poi: ApachePoiAdapter, monkeypatch: pytest.MonkeyPatch, tmp_path: Path
    ) -> None:
        class MissingTool:
            name = "apache-poi"
            command = ("missing-poi-helper",)

            def is_available(self) -> bool:
                return False

            def resolve_executable(self) -> None:
                return None

            def resolved_command(self) -> None:
                return None

        monkeypatch.setattr(ApachePoiAdapter, "tool", classmethod(lambda cls: MissingTool()))
        wb = poi.create_workbook()
        poi.add_sheet(wb, "S")
        poi.write_cell_value(wb, "S", "A1", CellValue(type=CellType.STRING, value="x"))
        with pytest.raises(FileNotFoundError):
            poi.save_workbook(wb, tmp_path / "missing.xlsx")


@pytest.mark.skipif(
    not ApachePoiAdapter.is_available(), reason="Apache POI helper is not available"
)
class TestApachePoiWriteRoundtrip:
    def test_string_formula_and_multiple_sheets(
        self, poi: ApachePoiAdapter, opxl: OpenpyxlAdapter, tmp_path: Path
    ) -> None:
        path = tmp_path / "poi_multi.xlsx"
        wb = poi.create_workbook()
        poi.add_sheet(wb, "S1")
        poi.add_sheet(wb, "S2")
        poi.write_cell_value(wb, "S1", "A1", CellValue(type=CellType.STRING, value="hello"))
        poi.write_cell_value(wb, "S1", "B1", CellValue(type=CellType.NUMBER, value=42.5))
        poi.write_cell_value(
            wb,
            "S1",
            "C1",
            CellValue(type=CellType.FORMULA, value="=B1*2", formula="=B1*2"),
        )
        poi.write_cell_value(wb, "S2", "A1", CellValue(type=CellType.BOOLEAN, value=True))
        poi.save_workbook(wb, path)

        rb = opxl.open_workbook(path)
        assert opxl.read_cell_value(rb, "S1", "A1").value == "hello"
        assert opxl.read_cell_value(rb, "S1", "B1").value == 42.5
        assert opxl.read_cell_value(rb, "S1", "C1").type == CellType.FORMULA
        assert opxl.read_cell_value(rb, "S2", "A1").value is True
        opxl.close_workbook(rb)

    def test_format_merge_comment_hyperlink_freeze_and_validation(
        self, poi: ApachePoiAdapter, tmp_path: Path
    ) -> None:
        import openpyxl

        path = tmp_path / "poi_features.xlsx"
        wb = poi.create_workbook()
        poi.add_sheet(wb, "S")
        poi.write_cell_value(wb, "S", "A1", CellValue(type=CellType.STRING, value="bold"))
        poi.write_cell_format(wb, "S", "A1", CellFormat(bold=True))
        poi.write_cell_value(wb, "S", "B2", CellValue(type=CellType.STRING, value="link"))
        poi.add_hyperlink(
            wb, "S", {"cell": "B2", "target": "https://poi.apache.org/", "label": "POI"}
        )
        poi.add_comment(wb, "S", {"cell": "C3", "text": "note", "author": "ExcelBench"})
        poi.merge_cells(wb, "S", "D1:E1")
        poi.set_freeze_panes(wb, "S", {"x_split": 1, "y_split": 1})
        poi.add_data_validation(wb, "S", {"cell": "F2", "values": ["Open", "Closed"]})
        poi.save_workbook(wb, path)

        rb = openpyxl.load_workbook(path)
        ws = rb["S"]
        assert ws["A1"].font.bold is True
        assert ws["B2"].hyperlink is not None
        assert ws["C3"].comment is not None
        assert "D1:E1" in {str(rng) for rng in ws.merged_cells.ranges}
        assert ws.freeze_panes == "B2"
        assert len(list(ws.data_validations.dataValidation)) == 1
        rb.close()

    def test_image_roundtrip(self, poi: ApachePoiAdapter, tmp_path: Path) -> None:
        import openpyxl

        path = tmp_path / "poi_image.xlsx"
        wb = poi.create_workbook()
        poi.add_sheet(wb, "S")
        poi.add_image(
            wb, "S", {"cell": "B2", "path": "fixtures/images/sample.png", "anchor": "oneCell"}
        )
        poi.save_workbook(wb, path)

        rb = openpyxl.load_workbook(path)
        ws = rb["S"]
        assert len(getattr(ws, "_images", [])) == 1
        rb.close()

    def test_split_pane_roundtrip(
        self, poi: ApachePoiAdapter, opxl: OpenpyxlAdapter, tmp_path: Path
    ) -> None:
        path = tmp_path / "poi_split.xlsx"
        wb = poi.create_workbook()
        poi.add_sheet(wb, "Split")
        poi.set_freeze_panes(wb, "Split", {"freeze": {"mode": "split", "x_split": 1, "y_split": 2}})
        poi.save_workbook(wb, path)

        rb = opxl.open_workbook(path)
        actual = opxl.read_freeze_panes(rb, "Split")
        assert actual["mode"] == "split"
        assert actual["x_split"] == 1
        assert actual["y_split"] == 2
        opxl.close_workbook(rb)

    def test_conditional_format_roundtrip(
        self, poi: ApachePoiAdapter, opxl: OpenpyxlAdapter, tmp_path: Path
    ) -> None:
        path = tmp_path / "poi_cf.xlsx"
        wb = poi.create_workbook()
        poi.add_sheet(wb, "S")
        for row in range(2, 7):
            poi.write_cell_value(wb, "S", f"B{row}", CellValue(type=CellType.NUMBER, value=row - 1))
        poi.add_conditional_format(
            wb,
            "S",
            {
                "cf_rule": {
                    "range": "B2:B6",
                    "rule_type": "cellIs",
                    "operator": "greaterThan",
                    "formula": "5",
                    "format": {"bg_color": "#FFFF00"},
                }
            },
        )
        poi.save_workbook(wb, path)

        rb = opxl.open_workbook(path)
        rules = opxl.read_conditional_formats(rb, "S")
        assert any(
            rule.get("rule_type") == "cellIs" and rule.get("formula") == "5" for rule in rules
        )
        opxl.close_workbook(rb)

    def test_named_range_roundtrip(
        self, poi: ApachePoiAdapter, opxl: OpenpyxlAdapter, tmp_path: Path
    ) -> None:
        path = tmp_path / "poi_names.xlsx"
        wb = poi.create_workbook()
        poi.add_sheet(wb, "named_ranges")
        poi.add_sheet(wb, "Targets")
        poi.write_cell_value(wb, "named_ranges", "B2", CellValue(type=CellType.NUMBER, value=42))
        poi.write_cell_value(wb, "Targets", "A1", CellValue(type=CellType.STRING, value="Target"))
        poi.add_named_range(
            wb,
            "named_ranges",
            {
                "named_range": {
                    "name": "SingleCell",
                    "scope": "workbook",
                    "refers_to": "named_ranges!$B$2",
                }
            },
        )
        poi.add_named_range(
            wb,
            "named_ranges",
            {
                "named_range": {
                    "name": "LocalName",
                    "scope": "sheet",
                    "refers_to": "named_ranges!$B$2",
                }
            },
        )
        poi.add_named_range(
            wb,
            "named_ranges",
            {
                "named_range": {
                    "name": "OtherSheet",
                    "scope": "workbook",
                    "refers_to": "Targets!$A$1",
                }
            },
        )
        poi.save_workbook(wb, path)

        rb = opxl.open_workbook(path)
        names = opxl.read_named_ranges(rb, "named_ranges")
        assert any(
            item.get("name") == "SingleCell" and item.get("scope") == "workbook" for item in names
        )
        assert any(
            item.get("name") == "LocalName" and item.get("scope") == "sheet" for item in names
        )
        assert any(
            item.get("name") == "OtherSheet" and item.get("refers_to") == "Targets!$A$1"
            for item in names
        )
        opxl.close_workbook(rb)

    def test_border_roundtrip(
        self, poi: ApachePoiAdapter, opxl: OpenpyxlAdapter, tmp_path: Path
    ) -> None:
        from excelbench.models import BorderEdge, BorderInfo, BorderStyle

        path = tmp_path / "poi_border.xlsx"
        wb = poi.create_workbook()
        poi.add_sheet(wb, "S")
        poi.write_cell_value(wb, "S", "B2", CellValue(type=CellType.STRING, value="Border"))
        poi.write_cell_border(
            wb,
            "S",
            "B2",
            BorderInfo(
                top=BorderEdge(style=BorderStyle.THIN, color="#FF0000"),
                bottom=BorderEdge(style=BorderStyle.DOUBLE, color="#00FF00"),
                left=BorderEdge(style=BorderStyle.MEDIUM, color="#0000FF"),
                right=BorderEdge(style=BorderStyle.DASHED, color="#FFFF00"),
            ),
        )
        poi.save_workbook(wb, path)

        rb = opxl.open_workbook(path)
        border = opxl.read_cell_border(rb, "S", "B2")
        assert (
            border.top
            and border.top.style.value == "thin"
            and border.top.color.upper() == "#FF0000"
        )
        assert (
            border.bottom
            and border.bottom.style.value == "double"
            and border.bottom.color.upper() == "#00FF00"
        )
        assert (
            border.left
            and border.left.style.value == "medium"
            and border.left.color.upper() == "#0000FF"
        )
        assert (
            border.right
            and border.right.style.value == "dashed"
            and border.right.color.upper() == "#FFFF00"
        )
        opxl.close_workbook(rb)

    def test_diagonal_border_roundtrip(
        self, poi: ApachePoiAdapter, opxl: OpenpyxlAdapter, tmp_path: Path
    ) -> None:
        from excelbench.models import BorderEdge, BorderInfo, BorderStyle

        path = tmp_path / "poi_diag_border.xlsx"
        wb = poi.create_workbook()
        poi.add_sheet(wb, "S")
        poi.write_cell_value(wb, "S", "B2", CellValue(type=CellType.STRING, value="Diag"))
        poi.write_cell_border(
            wb,
            "S",
            "B2",
            BorderInfo(
                diagonal_up=BorderEdge(style=BorderStyle.THIN, color="#FF0000"),
                diagonal_down=BorderEdge(style=BorderStyle.THIN, color="#0000FF"),
            ),
        )
        poi.save_workbook(wb, path)

        rb = opxl.open_workbook(path)
        border = opxl.read_cell_border(rb, "S", "B2")
        assert border.diagonal_up and border.diagonal_up.style.value == "thin"
        assert border.diagonal_down and border.diagonal_down.style.value == "thin"
        opxl.close_workbook(rb)

    def test_stop_if_true_roundtrip(
        self, poi: ApachePoiAdapter, opxl: OpenpyxlAdapter, tmp_path: Path
    ) -> None:
        path = tmp_path / "poi_cf_stop.xlsx"
        wb = poi.create_workbook()
        poi.add_sheet(wb, "S")
        for row in range(7, 10):
            poi.write_cell_value(wb, "S", f"B{row}", CellValue(type=CellType.NUMBER, value=row - 6))
        poi.add_conditional_format(
            wb,
            "S",
            {
                "cf_rule": {
                    "range": "B7:B9",
                    "rule_type": "cellIs",
                    "operator": "lessThan",
                    "formula": "3",
                    "stop_if_true": True,
                    "format": {"bg_color": "#FF0000"},
                }
            },
        )
        poi.save_workbook(wb, path)

        rb = opxl.open_workbook(path)
        rules = opxl.read_conditional_formats(rb, "S")
        assert any(
            rule.get("rule_type") == "cellIs" and rule.get("stop_if_true") is True for rule in rules
        )
        opxl.close_workbook(rb)

    def test_text_format_roundtrip(
        self, poi: ApachePoiAdapter, opxl: OpenpyxlAdapter, tmp_path: Path
    ) -> None:
        path = tmp_path / "poi_text_format.xlsx"
        wb = poi.create_workbook()
        poi.add_sheet(wb, "S")
        poi.write_cell_value(wb, "S", "B2", CellValue(type=CellType.STRING, value="Format"))
        poi.write_cell_format(
            wb,
            "S",
            "B2",
            CellFormat(
                bold=True,
                italic=True,
                underline="double",
                strikethrough=True,
                font_name="Arial",
                font_size=16,
                font_color="#FF0000",
            ),
        )
        poi.save_workbook(wb, path)

        rb = opxl.open_workbook(path)
        fmt = opxl.read_cell_format(rb, "S", "B2")
        assert fmt.bold is True
        assert fmt.italic is True
        assert fmt.underline == "double"
        assert fmt.strikethrough is True
        assert fmt.font_name == "Arial"
        assert float(fmt.font_size or 0) == 16.0
        assert fmt.font_color and fmt.font_color.upper() == "#FF0000"
        opxl.close_workbook(rb)

    def test_alignment_roundtrip(
        self, poi: ApachePoiAdapter, opxl: OpenpyxlAdapter, tmp_path: Path
    ) -> None:
        path = tmp_path / "poi_alignment.xlsx"
        wb = poi.create_workbook()
        poi.add_sheet(wb, "S")
        poi.write_cell_value(wb, "S", "B2", CellValue(type=CellType.STRING, value="Line 1\nLine 2"))
        poi.write_cell_format(
            wb,
            "S",
            "B2",
            CellFormat(h_align="center", v_align="bottom", wrap=True, rotation=45, indent=2),
        )
        poi.save_workbook(wb, path)

        rb = opxl.open_workbook(path)
        fmt = opxl.read_cell_format(rb, "S", "B2")
        assert fmt.h_align == "center"
        assert fmt.v_align == "bottom"
        assert fmt.wrap is True
        assert fmt.rotation == 45
        assert fmt.indent == 2
        opxl.close_workbook(rb)
