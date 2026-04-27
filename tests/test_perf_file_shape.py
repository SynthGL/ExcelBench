"""Tests for the Sprint 3 file-shape extension.

Covers the new ``n_sheets`` / ``sheet_pattern`` workload fields in
``_run_workload_write`` and ``_run_workload_read``, plus the CLI helpers
``_resolve_file_shape_features`` / ``_file_shape_fixtures_stale`` and the
dashboard helper ``_section_file_shape``.

These tests use small fixtures (~100-cell sheets) so they finish in <100ms
each. Larger-tier behavior is exercised end-to-end via the smoke runs
documented in DEC-020.
"""

from __future__ import annotations

import json
from datetime import UTC, datetime
from pathlib import Path

import pytest
from openpyxl import Workbook

from excelbench.generator.generate import write_manifest
from excelbench.harness.adapters.openpyxl_adapter import OpenpyxlAdapter
from excelbench.models import Importance, Manifest
from excelbench.models import TestCase as BenchCase
from excelbench.models import TestFile as BenchFile
from excelbench.perf.runner import run_perf


def _build_many_sheets_suite(
    tmp_path: Path,
    *,
    scenario: str,
    n_sheets: int,
    rows: int,
    cols: int,
) -> Path:
    """Synthesize a fixture with N sheets named Sheet1..SheetN.

    Each sheet has the same rows×cols int grid. The bulk_read workload
    cycles through all N sheets via sheet_pattern.
    """
    suite = tmp_path / "suite"
    suite.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Drop the default Sheet so the workbook has only the named ones.
    if wb.sheetnames:
        default = wb.active
        if default is not None:
            wb.remove(default)
    for i in range(1, n_sheets + 1):
        ws = wb.create_sheet(f"Sheet{i}")
        for r in range(1, rows + 1):
            for c in range(1, cols + 1):
                ws.cell(row=r, column=c, value=r * cols + c)

    (suite / "tier0").mkdir(parents=True, exist_ok=True)
    wb_path = suite / "tier0" / f"00_{scenario}.xlsx"
    wb.save(wb_path)

    end_col_letter = chr(ord("A") + cols - 1)
    rng = f"A1:{end_col_letter}{rows}"

    files = []
    # Read workload: fan out across N sheets.
    read_feature = f"file_shape_{scenario}_bulk_read"
    files.append(
        BenchFile(
            path=f"tier0/00_{scenario}.xlsx",
            feature=read_feature,
            tier=0,
            file_format="xlsx",
            test_cases=[
                BenchCase(
                    id=read_feature,
                    label=read_feature,
                    row=1,
                    expected={
                        "workload": {
                            "scenario": read_feature,
                            "op": "bulk_sheet_values",
                            "operations": ["read"],
                            "sheet": "Sheet1",
                            "range": rng,
                            "n_sheets": n_sheets,
                            "sheet_pattern": "Sheet{i}",
                        }
                    },
                    importance=Importance.BASIC,
                )
            ],
        )
    )

    # Write workload: same fan-out.
    write_feature = f"file_shape_{scenario}_bulk_write"
    files.append(
        BenchFile(
            path=f"tier0/00_{scenario}.xlsx",
            feature=write_feature,
            tier=0,
            file_format="xlsx",
            test_cases=[
                BenchCase(
                    id=write_feature,
                    label=write_feature,
                    row=1,
                    expected={
                        "workload": {
                            "scenario": write_feature,
                            "op": "bulk_write_grid",
                            "operations": ["write"],
                            "sheet": "Sheet1",
                            "range": rng,
                            "value_type": "number",
                            "n_sheets": n_sheets,
                            "sheet_pattern": "Sheet{i}",
                        }
                    },
                    importance=Importance.BASIC,
                )
            ],
        )
    )

    manifest = Manifest(
        generated_at=datetime.now(UTC),
        excel_version="test",
        generator_version="test",
        file_format="xlsx",
        files=files,
    )
    write_manifest(manifest, suite / "manifest.json")
    return suite


# ---------------------------------------------------------------------------
# Runner: n_sheets fan-out
# ---------------------------------------------------------------------------


def test_many_sheets_read_op_count_scales_with_n_sheets(tmp_path: Path) -> None:
    """Read workload over 5 sheets × 4 cells each = 20 op_count."""
    suite = _build_many_sheets_suite(
        tmp_path, scenario="ms_5x4", n_sheets=5, rows=2, cols=2
    )
    results = run_perf(
        suite,
        adapters=[OpenpyxlAdapter()],
        warmup=0,
        iters=1,
        breakdown=False,
        features=["file_shape_ms_5x4_bulk_read"],
    )
    row = results.results[0]
    read = row.perf["read"]
    assert read is not None
    # 2*2 cells × 5 sheets = 20.
    assert read.op_count == 20


def test_many_sheets_write_op_count_scales_with_n_sheets(tmp_path: Path) -> None:
    """Write workload over 3 sheets × 4 cells each = 12 op_count."""
    suite = _build_many_sheets_suite(
        tmp_path, scenario="ms_3x4w", n_sheets=3, rows=2, cols=2
    )
    results = run_perf(
        suite,
        adapters=[OpenpyxlAdapter()],
        warmup=0,
        iters=1,
        breakdown=False,
        features=["file_shape_ms_3x4w_bulk_write"],
    )
    row = results.results[0]
    write = row.perf["write"]
    assert write is not None
    assert write.op_count == 12


def test_many_sheets_default_sheet_pattern(tmp_path: Path) -> None:
    """Default sheet_pattern 'Sheet{i}' is used when not specified."""
    suite = tmp_path / "suite"
    suite.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    if wb.sheetnames:
        default = wb.active
        if default is not None:
            wb.remove(default)
    for i in range(1, 4):
        ws = wb.create_sheet(f"Sheet{i}")
        ws["A1"] = i
        ws["A2"] = i * 10

    (suite / "tier0").mkdir(parents=True, exist_ok=True)
    wb.save(suite / "tier0" / "00_default_pattern.xlsx")

    feature = "file_shape_default_pattern_bulk_read"
    files = [
        BenchFile(
            path="tier0/00_default_pattern.xlsx",
            feature=feature,
            tier=0,
            file_format="xlsx",
            test_cases=[
                BenchCase(
                    id=feature,
                    label=feature,
                    row=1,
                    expected={
                        "workload": {
                            "scenario": feature,
                            "op": "bulk_sheet_values",
                            "operations": ["read"],
                            "sheet": "Sheet1",
                            "range": "A1:A2",
                            "n_sheets": 3,
                            # No sheet_pattern — should default to "Sheet{i}".
                        }
                    },
                    importance=Importance.BASIC,
                )
            ],
        )
    ]
    write_manifest(
        Manifest(
            generated_at=datetime.now(UTC),
            excel_version="test",
            generator_version="test",
            file_format="xlsx",
            files=files,
        ),
        suite / "manifest.json",
    )

    results = run_perf(
        suite,
        adapters=[OpenpyxlAdapter()],
        warmup=0,
        iters=1,
        breakdown=False,
    )
    read = results.results[0].perf["read"]
    assert read is not None
    assert read.op_count == 6  # 2 cells × 3 sheets


# ---------------------------------------------------------------------------
# CLI: _resolve_file_shape_features
# ---------------------------------------------------------------------------


def test_resolve_file_shape_features_default_all() -> None:
    from excelbench.cli import _resolve_file_shape_features

    reads, writes, labels = _resolve_file_shape_features(
        shapes_arg="all", rows=100_000
    )
    assert all(r.startswith("file_shape_") and r.endswith("_bulk_read") for r in reads)
    assert all(w.startswith("file_shape_") and w.endswith("_bulk_write") for w in writes)
    # 100k cap: 4 categories × at-most-2 tiers visible (10k + 100k tiers).
    # wide_10k, wide_100k, tall_10k, tall_100k, sparse_10pct_10k,
    # sparse_10pct_100k, many_sheets_10x10k → 7 labels.
    assert "wide_10k" in labels
    assert "wide_100k" in labels
    assert "tall_100k" in labels
    assert "many_sheets_10x10k" in labels
    # 1M scenarios excluded.
    assert "wide_1m" not in labels
    assert "many_sheets_100x10k" not in labels


def test_resolve_file_shape_features_filter_by_category() -> None:
    from excelbench.cli import _resolve_file_shape_features

    _, _, labels = _resolve_file_shape_features(shapes_arg="wide", rows=1_000_000)
    assert all(label.startswith("wide_") for label in labels)
    assert "tall_10k" not in labels
    assert "many_sheets_10x10k" not in labels


def test_resolve_file_shape_features_unknown_category_raises() -> None:
    from excelbench.cli import _resolve_file_shape_features

    with pytest.raises(ValueError, match="Unknown --shapes"):
        _resolve_file_shape_features(shapes_arg="not_a_shape", rows=1_000_000)


def test_resolve_file_shape_features_rows_too_low_raises() -> None:
    from excelbench.cli import _resolve_file_shape_features

    with pytest.raises(ValueError, match="No file-shape scenarios match"):
        _resolve_file_shape_features(shapes_arg="all", rows=500)


def test_resolve_file_shape_features_includes_1m_when_requested() -> None:
    from excelbench.cli import _resolve_file_shape_features

    _, _, labels = _resolve_file_shape_features(
        shapes_arg="all", rows=1_000_000
    )
    assert "wide_1m" in labels
    assert "tall_1m" in labels
    assert "sparse_10pct_1m" in labels
    assert "many_sheets_100x10k" in labels
    assert "many_sheets_1000x1k" in labels


# ---------------------------------------------------------------------------
# CLI: _file_shape_fixtures_stale
# ---------------------------------------------------------------------------


def test_file_shape_fixtures_stale_missing_manifest(tmp_path: Path) -> None:
    from excelbench.cli import _file_shape_fixtures_stale

    manifest = tmp_path / "manifest.json"
    generator = tmp_path / "gen.py"
    generator.write_text("# generator")
    assert _file_shape_fixtures_stale(manifest, generator, needs_1m=False) is True


def test_file_shape_fixtures_stale_generator_newer(tmp_path: Path) -> None:
    import os

    from excelbench.cli import _file_shape_fixtures_stale

    manifest = tmp_path / "manifest.json"
    manifest.write_text('{"files": []}')
    generator = tmp_path / "gen.py"
    generator.write_text("# generator")
    later = manifest.stat().st_mtime + 5.0
    os.utime(generator, (later, later))
    assert _file_shape_fixtures_stale(manifest, generator, needs_1m=False) is True


def test_file_shape_fixtures_stale_no_1m_in_manifest(tmp_path: Path) -> None:
    from excelbench.cli import _file_shape_fixtures_stale

    manifest = tmp_path / "manifest.json"
    manifest.write_text(
        json.dumps(
            {
                "files": [
                    {"feature": "file_shape_wide_10k_bulk_read", "path": "x.xlsx"}
                ]
            }
        )
    )
    generator = tmp_path / "gen.py"
    generator.write_text("# generator")
    assert _file_shape_fixtures_stale(manifest, generator, needs_1m=True) is True


def test_file_shape_fixtures_stale_fresh_manifest_with_1m(tmp_path: Path) -> None:
    import os

    from excelbench.cli import _file_shape_fixtures_stale

    manifest = tmp_path / "manifest.json"
    manifest.write_text(
        json.dumps(
            {
                "files": [
                    {"feature": "file_shape_wide_1m_bulk_read", "path": "x.xlsx"}
                ]
            }
        )
    )
    generator = tmp_path / "gen.py"
    generator.write_text("# generator")
    later = generator.stat().st_mtime + 5.0
    os.utime(manifest, (later, later))
    assert _file_shape_fixtures_stale(manifest, generator, needs_1m=True) is False
    assert _file_shape_fixtures_stale(manifest, generator, needs_1m=False) is False


def test_file_shape_fixtures_stale_recognizes_1m_via_many_sheets(tmp_path: Path) -> None:
    """100x10k and 1000x1k satisfy the 1M gate via their label patterns."""
    import os

    from excelbench.cli import _file_shape_fixtures_stale

    manifest = tmp_path / "manifest.json"
    manifest.write_text(
        json.dumps(
            {
                "files": [
                    {"feature": "file_shape_many_sheets_100x10k_bulk_read", "path": "x.xlsx"}
                ]
            }
        )
    )
    generator = tmp_path / "gen.py"
    generator.write_text("# generator")
    later = generator.stat().st_mtime + 5.0
    os.utime(manifest, (later, later))
    assert _file_shape_fixtures_stale(manifest, generator, needs_1m=True) is False


def test_file_shape_fixtures_stale_data_shape_only_manifest_is_stale(tmp_path: Path) -> None:
    """Cross-command guard: a manifest written by perf-shape (data_shape only)
    must be treated as stale by perf-file-shape, regardless of needs_1m or
    fresh mtime, so the file-shape fixtures get regenerated.
    """
    import os

    from excelbench.cli import _file_shape_fixtures_stale

    manifest = tmp_path / "manifest.json"
    manifest.write_text(
        json.dumps(
            {
                "files": [
                    {"feature": "data_shape_int_10k_bulk_read", "path": "x.xlsx"},
                    {"feature": "data_shape_int_10k_bulk_write", "path": "x.xlsx"},
                ]
            }
        )
    )
    generator = tmp_path / "gen.py"
    generator.write_text("# generator")
    later = generator.stat().st_mtime + 5.0
    os.utime(manifest, (later, later))
    # Both flag values must report stale: the manifest contains zero
    # file_shape_* entries even though it's fresh on disk.
    assert _file_shape_fixtures_stale(manifest, generator, needs_1m=False) is True
    assert _file_shape_fixtures_stale(manifest, generator, needs_1m=True) is True


# ---------------------------------------------------------------------------
# Dashboard: _section_file_shape
# ---------------------------------------------------------------------------


def _file_shape_perf_payload(
    *, libs: list[str], scenarios: list[tuple[str, int]]
) -> dict[str, object]:
    """Synthesize a perf dict for dashboard tests.

    ``scenarios`` is a list of (label, op_count) pairs. Each library gets one
    bulk_read + one bulk_write entry per scenario.
    """
    results = []
    for lib in libs:
        for label, op_count in scenarios:
            for op in ("read", "write"):
                feature = f"file_shape_{label}_bulk_{op}"
                # Vary p50 by lib + label for realistic heatmap shape.
                base = 10.0 + len(lib) * 2 + len(label)
                results.append(
                    {
                        "library": lib,
                        "feature": feature,
                        "perf": {
                            op: {
                                "wall_ms": {"p50": base, "p95": base * 1.2},
                                "op_count": op_count,
                                "op_unit": "cells",
                            }
                        },
                    }
                )
    return {"results": results}


def test_section_file_shape_returns_empty_for_no_perf() -> None:
    from excelbench.results.html_dashboard import _section_file_shape

    assert _section_file_shape(None) == ""
    assert _section_file_shape({}) == ""
    assert _section_file_shape({"results": []}) == ""


def test_section_file_shape_renders_heatmap_with_categories() -> None:
    from excelbench.results.html_dashboard import _section_file_shape

    perf = _file_shape_perf_payload(
        libs=["wolfxl", "openpyxl"],
        scenarios=[
            ("wide_10k", 10_000),
            ("tall_10k", 10_000),
            ("sparse_10pct_10k", 1_000),
            ("many_sheets_10x10k", 100_000),
        ],
    )
    html = _section_file_shape(perf)
    assert html  # non-empty
    assert 'id="file-shape"' in html
    assert "wolfxl" in html
    assert "openpyxl" in html
    # Categories appear as column headers.
    assert ">wide<" in html
    assert ">tall<" in html
    assert ">sparse<" in html
    assert ">many_sheets<" in html
    # Both read + write tables.
    assert "Read" in html
    assert "Write" in html


def test_section_file_shape_skips_features_without_op_count() -> None:
    """An entry without op_count is dropped silently — heatmap stays sane."""
    from excelbench.results.html_dashboard import _section_file_shape

    perf = {
        "results": [
            {
                "library": "wolfxl",
                "feature": "file_shape_wide_10k_bulk_read",
                "perf": {
                    "read": {"wall_ms": {"p50": 1.0}, "op_count": 10_000}
                },
            },
            {
                "library": "broken",
                "feature": "file_shape_wide_10k_bulk_read",
                "perf": {"read": {"wall_ms": {"p50": 1.0}}},  # no op_count
            },
        ]
    }
    html = _section_file_shape(perf)
    assert "wolfxl" in html
    assert "broken" not in html


def test_file_shape_category_helper() -> None:
    from excelbench.results.html_dashboard import _file_shape_category_for_label

    assert _file_shape_category_for_label("wide_10k") == "wide"
    assert _file_shape_category_for_label("tall_1m") == "tall"
    assert _file_shape_category_for_label("sparse_10pct_100k") == "sparse"
    assert _file_shape_category_for_label("many_sheets_10x10k") == "many_sheets"
    assert _file_shape_category_for_label("not_a_shape") is None


# ---------------------------------------------------------------------------
# perf-file-shape CLI command
# ---------------------------------------------------------------------------


def _build_file_shape_fixture_manifest(
    fixtures_dir: Path, *, label: str, n_sheets: int = 1
) -> None:
    """Synthesize a minimal file_shape_* manifest the CLI can consume.

    Skips xlsxwriter generation by writing one fake bulk_read + bulk_write
    feature with a tiny range. Manifest mtime is fresh so the CLI sees it
    as not stale.
    """
    fixtures_dir.mkdir(parents=True, exist_ok=True)
    tier_dir = fixtures_dir / "tier0"
    tier_dir.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    if wb.sheetnames:
        default = wb.active
        if default is not None:
            wb.remove(default)
    if n_sheets <= 1:
        ws = wb.create_sheet("S1")
        ws["A1"] = 1
        ws["B1"] = 2
    else:
        for i in range(1, n_sheets + 1):
            ws = wb.create_sheet(f"Sheet{i}")
            ws["A1"] = i
            ws["B1"] = i * 10

    wb_path = tier_dir / f"00_file_shape_{label}.xlsx"
    wb.save(wb_path)

    files = []
    sheet_name = "S1" if n_sheets <= 1 else "Sheet1"
    for op in ("read", "write"):
        feature = f"file_shape_{label}_bulk_{op}"
        workload: dict[str, object] = {
            "scenario": feature,
            "op": "bulk_write_grid" if op == "write" else "bulk_sheet_values",
            "operations": [op],
            "sheet": sheet_name,
            "range": "A1:B1",
        }
        if op == "write":
            workload["value_type"] = "number"
        if n_sheets > 1:
            workload["n_sheets"] = n_sheets
            workload["sheet_pattern"] = "Sheet{i}"

        files.append(
            BenchFile(
                path=f"tier0/00_file_shape_{label}.xlsx",
                feature=feature,
                tier=0,
                file_format="xlsx",
                test_cases=[
                    BenchCase(
                        id=feature,
                        label=feature,
                        row=1,
                        expected={"workload": workload},
                        importance=Importance.BASIC,
                    )
                ],
            )
        )

    write_manifest(
        Manifest(
            generated_at=datetime.now(UTC),
            excel_version="test",
            generator_version="test",
            file_format="xlsx",
            files=files,
        ),
        fixtures_dir / "manifest.json",
    )


def test_perf_file_shape_invalid_memory_mode_exits(tmp_path: Path) -> None:
    import typer

    from excelbench.cli import perf_file_shape

    with pytest.raises(typer.Exit) as exc:
        perf_file_shape(
            rows=10_000,
            shapes="wide",
            output=tmp_path / "out",
            fixtures=tmp_path / "fixtures",
            adapters=None,
            warmup=0,
            iters=1,
            iteration_policy="fixed",
            breakdown=False,
            memory_mode="not_a_real_mode",
            regenerate=False,
        )
    assert exc.value.exit_code == 1


def test_perf_file_shape_invalid_shape_exits(tmp_path: Path) -> None:
    import typer

    from excelbench.cli import perf_file_shape

    with pytest.raises(typer.Exit) as exc:
        perf_file_shape(
            rows=10_000,
            shapes="not_a_shape",
            output=tmp_path / "out",
            fixtures=tmp_path / "fixtures",
            adapters=None,
            warmup=0,
            iters=1,
            iteration_policy="fixed",
            breakdown=False,
            memory_mode="getrusage",
            regenerate=False,
        )
    assert exc.value.exit_code == 1


def test_perf_file_shape_unknown_adapter_exits(tmp_path: Path) -> None:
    import typer

    from excelbench.cli import perf_file_shape

    _build_file_shape_fixture_manifest(tmp_path / "fixtures", label="wide_10k")

    with pytest.raises(typer.Exit) as exc:
        perf_file_shape(
            rows=10_000,
            shapes="wide",
            output=tmp_path / "out",
            fixtures=tmp_path / "fixtures",
            adapters=["nonexistent_lib_xyz"],
            warmup=0,
            iters=1,
            iteration_policy="fixed",
            breakdown=False,
            memory_mode="getrusage",
            regenerate=False,
        )
    assert exc.value.exit_code == 1


def test_perf_file_shape_happy_path_writes_outputs(tmp_path: Path) -> None:
    """End-to-end: perf_file_shape walks the manifest and writes results."""
    from excelbench.cli import perf_file_shape

    fixtures = tmp_path / "fixtures"
    output = tmp_path / "out"
    _build_file_shape_fixture_manifest(fixtures, label="wide_10k")

    perf_file_shape(
        rows=10_000,
        shapes="wide",
        output=output,
        fixtures=fixtures,
        adapters=["openpyxl"],
        warmup=0,
        iters=1,
        iteration_policy="fixed",
        breakdown=False,
        memory_mode="getrusage",
        regenerate=False,
    )

    assert (output / "perf" / "results.json").exists()
    assert (output / "perf" / "matrix.csv").exists()
    assert (output / "perf" / "history.jsonl").exists()
